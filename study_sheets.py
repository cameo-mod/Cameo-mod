import openpyxl

wb = openpyxl.load_workbook('docs/design/cameo_armor_system.xlsx', data_only=False)

# Read Weapon Types sheet
ws = wb['Weapon Types']
print(f"=== Weapon Types sheet: {ws.max_row} rows x {ws.max_column} cols ===\n")
for row in range(1, ws.max_row + 1):
    row_data = []
    has_content = False
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            has_content = True
            row_data.append(f"{cell.coordinate}={cell.value}")
    if has_content:
        print(f"  Row {row}: {' | '.join(row_data)}")

# Read Armor Types sheet
ws2 = wb['Armor Types']
print(f"\n=== Armor Types sheet: {ws2.max_row} rows x {ws2.max_column} cols ===\n")
for row in range(1, ws2.max_row + 1):
    row_data = []
    has_content = False
    for col in range(1, ws2.max_column + 1):
        cell = ws2.cell(row=row, column=col)
        if cell.value is not None:
            has_content = True
            row_data.append(f"{cell.coordinate}={cell.value}")
    if has_content:
        print(f"  Row {row}: {' | '.join(row_data)}")

# Check how existing tabs handle multi-weapon units - look at GDI Battle Tank
ws3 = wb['Tanks']
print("\n=== Tanks: GDI Battle Tank (multi-weapon example) ===")
for row in range(2, ws3.max_row + 1):
    name = ws3.cell(row=row, column=2).value
    if name and 'Battle Tank' in str(name):
        for col in range(1, ws3.max_column + 1):
            cell = ws3.cell(row=row, column=col)
            if cell.value is not None:
                print(f"  {cell.coordinate}: {cell.value}")
        break

# Check a few more multi-weapon examples in Vehicles
ws4 = wb['Vehicles']
print("\n=== Vehicles: first 10 rows ===")
for row in range(2, min(12, ws4.max_row + 1)):
    name = ws4.cell(row=row, column=2).value
    if name:
        g = ws4.cell(row=row, column=7).value
        h = ws4.cell(row=row, column=8).value
        i = ws4.cell(row=row, column=9).value
        print(f"  Row {row}: {name} | Dmg={g} | WC={h} | Reload={i}")

# Check Aircraft tab for unarmed/transport units
ws5 = wb['Aircraft']
print("\n=== Aircraft: first 15 rows ===")
for row in range(2, min(17, ws5.max_row + 1)):
    name = ws5.cell(row=row, column=2).value
    if name:
        d = ws5.cell(row=row, column=4).value
        e = ws5.cell(row=row, column=5).value
        g = ws5.cell(row=row, column=7).value
        h = ws5.cell(row=row, column=8).value
        i = ws5.cell(row=row, column=9).value
        s = ws5.cell(row=row, column=19).value
        print(f"  Row {row}: {name} | HP={d} | Spd={e} | Dmg={g} | WC={h} | Rld={i} | Cost={s}")
