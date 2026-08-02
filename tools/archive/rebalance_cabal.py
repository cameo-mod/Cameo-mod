#!/usr/bin/env python3
"""
CABAL rebalance script:
1. Reads sheet inputs (HP, Speed, Damage, ReloadDelay, K, L, M, Cost)
2. Computes correct Range from the balance formula
3. Identifies broken formula references
4. Reports what needs to change in both sheet and yaml
"""
import openpyxl
import os

def compute_range(hp, speed, damage, reload, wclass, k, l, m, cost):
    """
    Solve for Range from the balance formula:
    Cost = (O + P + Q) / 3
    
    O = (HP/100000 + Speed/100 + Range*K/5 + DPS/200) * 200 * L * M
    P = (HP*Speed/25000 + Range*K*DPS/2.5) * L * M
    Q = HP*Speed*Range*K*DPS*L*M / 12500000
    
    DPS = Damage / ReloadDelay * WeaponClass
    
    Solving for Range:
    3*Cost/(L*M) = 200*(HP/100000 + Speed/100 + DPS/200) + HP*Speed/25000 
                   + Range*K*(40 + DPS/2.5 + HP*Speed*DPS/12500000)
    """
    dps = damage / reload * wclass
    h = hp / 100000
    s = speed / 100
    d = dps / 200
    p = hp * speed / 25000
    q_coeff = hp * speed * dps / 12500000
    
    # 3*Cost/(L*M) = 200*(h+s+d) + p + Range*K*(40 + DPS/2.5 + q_coeff)
    lhs = 3 * cost / (l * m)
    constant = 200 * (h + s + d) + p
    range_coeff = k * (40 + dps / 2.5 + q_coeff)
    
    range_val = (lhs - constant) / range_coeff
    return range_val, dps

def main():
    wb = openpyxl.load_workbook('docs/design/cameo_armor_system.xlsx', data_only=False)
    ws = wb['CABAL']
    
    print("=== CABAL Rebalance Analysis ===")
    print(f"{'Actor':<35} {'HP':>8} {'Spd':>5} {'Dmg':>8} {'Rld':>5} {'K':>4} {'L':>5} {'M':>4} {'Cost':>6} {'RngSheet':>10} {'RngCalc':>10} {'Match':>5} {'FormulaOK':>9}")
    print("-" * 140)
    
    issues = []
    
    for row in range(2, ws.max_row + 1):
        actor = ws.cell(row=row, column=3).value
        if not actor:
            continue
        
        name = ws.cell(row=row, column=2).value
        hp = ws.cell(row=row, column=4).value
        speed = ws.cell(row=row, column=5).value
        rng_sheet = ws.cell(row=row, column=6).value
        damage = ws.cell(row=row, column=7).value
        wclass = ws.cell(row=row, column=8).value
        reload = ws.cell(row=row, column=9).value
        k = ws.cell(row=row, column=11).value
        l = ws.cell(row=row, column=12).value
        m = ws.cell(row=row, column=13).value
        cost = ws.cell(row=row, column=19).value
        
        # Check formula references
        o_formula = ws.cell(row=row, column=15).value  # column O
        formula_ok = "YES"
        if o_formula and isinstance(o_formula, str) and o_formula.startswith('='):
            # Check if formula references the correct row
            expected = f"D{row}"
            if expected not in o_formula:
                formula_ok = f"BAD (refs wrong row)"
                issues.append(f"  Row {row} ({actor}): O formula references wrong row: {o_formula}")
        
        if damage is None or damage == 0 or reload is None or reload == 0:
            # No weapon — skip range computation
            print(f"{actor:<35} {hp:>8} {speed:>5} {damage or 0:>8} {reload or 0:>5} {k:>4} {l:>5} {m:>4} {cost:>6} {rng_sheet:>10} {'N/A':>10} {'N/A':>5} {formula_ok:>9}")
            continue
        
        rng_calc, dps = compute_range(hp, speed, damage, reload, wclass, k, l, m, cost)
        
        match = "OK" if abs(rng_sheet - rng_calc) < 0.01 else f"DIFF"
        
        print(f"{actor:<35} {hp:>8} {speed:>5} {damage:>8} {reload:>5} {k:>4} {l:>5} {m:>4} {cost:>6} {rng_sheet:>10.3f} {rng_calc:>10.3f} {match:>5} {formula_ok:>9}")
        
        if match != "OK":
            issues.append(f"  Row {row} ({actor}): Range sheet={rng_sheet:.3f} calc={rng_calc:.3f}")
    
    if issues:
        print(f"\n=== {len(issues)} ISSUES ===")
        for issue in issues:
            print(issue)
    else:
        print("\nNo issues found!")

if __name__ == "__main__":
    main()
