#!/usr/bin/env python3
"""
CABAL Full Rebalance Script
- Reads yaml weapons (with template inheritance for DamagesConcrete)
- Reads yaml rules (HP, Speed, Cost, weapons, prerequisites)
- Reads current sheet
- Computes correct K, L, M for each unit
- Computes correct sheet Damage and ReloadDelay from yaml
- Solves for Range from the balance formula
- Writes updated sheet
- Reports yaml changes needed
"""
import openpyxl
import re
import os
import math

MOD_ROOT = os.path.join(os.path.dirname(__file__), "..", "..", "mods", "cameo")
CABAL_DIR = os.path.join(MOD_ROOT, "ContentPacks", "TiberianSun", "CABAL")

# ============================================================
# 1. Read weapons from yaml (CABAL + shared)
# ============================================================

def read_weapons(fpath):
    """Read a weapons file and extract weapon definitions."""
    weapons = {}
    if not os.path.exists(fpath):
        return weapons
    with open(fpath, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    current = None
    for i, line in enumerate(lines):
        if re.match(r'^[A-Za-z\^][A-Za-z0-9_]*:', line) and not line.startswith(' '):
            current = line.split(':')[0].strip()
            if current not in weapons:
                weapons[current] = {
                    'range': None, 'reload': None, 'burst': 1,
                    'burst_delays': [], 'spread_damages': [],
                    'friendly_fire_damages': [], 'damages_concrete': 0,
                    'target_damages': [], 'inherits': [],
                }
        elif current:
            m = re.match(r'^\s+Inherits\S*:\s*(\S+)', line)
            if m:
                weapons[current]['inherits'].append(m.group(1))
            m = re.match(r'^\s+Range:\s*(\d+)', line)
            if m and weapons[current]['range'] is None:
                weapons[current]['range'] = int(m.group(1))
            m = re.match(r'^\s+ReloadDelay:\s*(\d+)', line)
            if m and weapons[current]['reload'] is None:
                weapons[current]['reload'] = int(m.group(1))
            m = re.match(r'^\s+Burst:\s*(\d+)', line)
            if m:
                weapons[current]['burst'] = int(m.group(1))
            m = re.match(r'^\s+BurstDelays:\s*(.+)', line)
            if m:
                weapons[current]['burst_delays'] = [int(x.strip()) for x in m.group(1).split(',')]
            
            # Warhead definitions
            m = re.match(r'^\s+Warhead@([^:]+):\s*(\S+)', line)
            if m:
                wh_name = m.group(1)
                wh_type = m.group(2)
                # Look ahead for Damage
                for j in range(i+1, min(i+5, len(lines))):
                    dm = re.match(r'^\s+Damage:\s*(\d+)', lines[j])
                    if dm:
                        dmg = int(dm.group(1))
                        if 'FriendlyFire' in wh_name:
                            weapons[current]['friendly_fire_damages'].append(dmg)
                        elif wh_type == 'SpreadDamage':
                            weapons[current]['spread_damages'].append(dmg)
                        elif wh_type == 'DamagesConcrete':
                            weapons[current]['damages_concrete'] += dmg
                        elif wh_type == 'TargetDamage':
                            weapons[current]['target_damages'].append(dmg)
                        break
    return weapons

def resolve_weapon(wname, all_weapons, depth=0):
    """Resolve a weapon's full stats including inherited template DamagesConcrete."""
    if wname not in all_weapons or depth > 10:
        return None
    w = all_weapons[wname]
    
    # Collect DamagesConcrete from inherits
    total_concrete = w['damages_concrete']
    for parent in w.get('inherits', []):
        if parent.startswith('^'):
            parent_w = resolve_weapon(parent, all_weapons, depth+1)
            if parent_w:
                total_concrete += parent_w['damages_concrete']
    
    return {
        'range': w['range'],
        'reload': w['reload'],
        'burst': w['burst'],
        'burst_delays': w['burst_delays'],
        'spread_damages': w['spread_damages'],
        'friendly_fire_damages': w['friendly_fire_damages'],
        'damages_concrete': total_concrete,
        'target_damages': w['target_damages'],
    }

# ============================================================
# 2. Read rules from yaml
# ============================================================

def read_rules():
    """Read CABAL rules files for actor stats and weapon mappings."""
    actors = {}
    for fname in ['infantry.yaml', 'vehicles.yaml', 'aircraft.yaml']:
        fpath = os.path.join(CABAL_DIR, 'rules', fname)
        if not os.path.exists(fpath):
            continue
        with open(fpath, 'r', encoding='utf-8') as f:
            content = f.read()
        
        current = None
        for line in content.split('\n'):
            if re.match(r'^[a-z][a-z0-9_]*:', line):
                current = line.split(':')[0].strip()
                actors[current] = {
                    'file': fname, 'weapons': [], 'prerequisites': [],
                    'inherits': [], 'hp': None, 'speed': None, 'cost': None,
                }
            elif current:
                m = re.match(r'^\s+Weapon:\s*(\S+)', line)
                if m:
                    actors[current]['weapons'].append(m.group(1))
                m = re.match(r'^\s+HP:\s*(\d+)', line)
                if m and actors[current]['hp'] is None:
                    actors[current]['hp'] = int(m.group(1))
                m = re.match(r'^\s+Speed:\s*(\d+)', line)
                if m and actors[current]['speed'] is None:
                    actors[current]['speed'] = int(m.group(1))
                m = re.match(r'^\s+Cost:\s*(\d+)', line)
                if m and actors[current]['cost'] is None:
                    actors[current]['cost'] = int(m.group(1))
                m = re.match(r'^\s+Prerequisites:\s*(.+)', line)
                if m:
                    actors[current]['prerequisites'] = [x.strip() for x in m.group(1).split(',')]
                m = re.match(r'^\s+Inherits\S*:\s*(\S+)', line)
                if m:
                    actors[current]['inherits'].append(m.group(1))
    return actors

# ============================================================
# 3. Determine K, L, M for each unit
# ============================================================

# L mapping from DESIGN.md
L_MAP = {
    # Infantry
    'HeavyInfantryTemplate': 0.8,
    'HeroInfantryTemplate': 1.0,
    'AntiTankAntiAirInfantryTemplate': 0.5,
    'LineBreakerTemplate': 0.8,  # Treated as Heavy Infantry
    'MechanicTemplate': 0.5,  # Support
    # Vehicles
    'MainBattleTankTemplate': 1.0,
    'HighTechTankTemplate': 1.0,
    'SupportVehicleTemplate': 1.25,
    'FireSupportTemplate': 1.0,
    'ArtilleryTemplate': 0.5,
    'EpicVehicleTemplate': 0.3,
    # Aircraft
    'HelicopterTemplate': 1.0,
    'SpaceshipTemplate': 1.0,
    'BomberTemplate': 1.0,
    'FlyingInfantryTemplate': 1.0,
    'UnarmedTransportHelicopterTemplate': 1.0,
}

def determine_L(actor_data):
    """Determine L from the unit's template inheritance."""
    for inh in actor_data.get('inherits', []):
        if inh in L_MAP:
            return L_MAP[inh]
    # Fallback: check for Epic
    for inh in actor_data.get('inherits', []):
        if 'Epic' in inh:
            return 0.3
    return 1.0  # Default

def determine_M(actor_data):
    """Determine M from tech tier based on prerequisites."""
    prereqs = actor_data.get('prerequisites', [])
    has_techcenter = any('techcenter' in p for p in prereqs)
    has_core = any('core' in p and 'cabal_core' in p for p in prereqs)
    has_radar = any('radar' in p for p in prereqs)
    has_helipad = any('helipad' in p for p in prereqs)
    
    # Epic units: M=1.0 regardless
    for inh in actor_data.get('inherits', []):
        if 'Epic' in inh:
            return 1.0
    
    if has_core:
        return 0.5  # T4/5
    if has_techcenter:
        return 0.75  # T3
    if has_radar or has_helipad:
        return 1.0  # T2 (M=1.0 for T1/T2)
    return 1.0  # T1

def determine_K(actor_data):
    """Determine K from special abilities."""
    k = 1.0
    inherits = actor_data.get('inherits', [])
    
    # Check for EMP
    if any('EMP' in inh for inh in inherits):
        k += 0.25
    if any('UnitDisable' in inh for inh in inherits):
        if not any('EMP' in inh for inh in inherits):
            k += 0.25
    
    # Check for mind control
    if 'MindControllerCA' in str(actor_data):
        k += 0.25
    
    # Check for cloak/stealth
    if any('Stealth' in inh or 'Cloak' in inh for inh in inherits):
        k += 0.25
    
    # Check for trap weapon
    weapons = actor_data.get('weapons', [])
    if any('Trap' in w for w in weapons):
        k += 0.25
    
    # Check for PromotionUnitBuff (doesn't count for K)
    
    # Cap at 2.0
    return min(k, 2.0)

# ============================================================
# 4. Balance formula
# ============================================================

def compute_cost(hp, speed, rng, damage, reload, wclass, k, l, m):
    dps = damage / reload * wclass
    o = (hp/100000 + speed/100 + rng*k/5 + dps/200) * 200 * l * m
    p = (hp*speed/25000 + rng*k*dps/2.5) * l * m
    q = hp*speed*rng*k*dps*l*m / 12500000
    r = (o + p + q) / 3
    return r, dps

def solve_range(hp, speed, damage, reload, wclass, k, l, m, cost):
    dps = damage / reload * wclass
    h = hp / 100000
    s = speed / 100
    d = dps / 200
    p_const = hp * speed / 25000
    q_coeff = hp * speed * dps / 12500000
    
    lhs = 3 * cost / (l * m)
    constant = 200 * (h + s + d) + p_const
    range_coeff = k * (40 + dps / 2.5 + q_coeff)
    
    if range_coeff == 0:
        return None
    return (lhs - constant) / range_coeff

# ============================================================
# 5. Main
# ============================================================

def main():
    # Read all weapons
    all_weapons = {}
    for fpath in [
        os.path.join(CABAL_DIR, 'weapons', 'weapons.yaml'),
        os.path.join(MOD_ROOT, 'weapons', 'weapons.yaml'),
        os.path.join(MOD_ROOT, 'ContentPacks', 'TiberianSun', 'Shared', 'weapons', 'weapons.yaml'),
    ]:
        all_weapons.update(read_weapons(fpath))
    
    # Read rules
    actors = read_rules()
    
    # Read sheet
    wb = openpyxl.load_workbook('docs/design/cameo_armor_system.xlsx', data_only=False)
    ws = wb['CABAL']
    
    # CABAL actors in sheet order
    sheet_rows = {}
    for row in range(2, ws.max_row + 1):
        actor = ws.cell(row=row, column=3).value
        if actor:
            sheet_rows[actor] = row
    
    print("=== CABAL FULL REBALANCE ANALYSIS ===")
    print()
    
    changes = []
    
    for actor, row in sorted(sheet_rows.items(), key=lambda x: x[1]):
        s_hp = ws.cell(row=row, column=4).value
        s_speed = ws.cell(row=row, column=5).value
        s_rng = ws.cell(row=row, column=6).value
        s_dmg = ws.cell(row=row, column=7).value
        s_wclass = ws.cell(row=row, column=8).value
        s_reload = ws.cell(row=row, column=9).value
        s_k = ws.cell(row=row, column=11).value
        s_l = ws.cell(row=row, column=12).value
        s_m = ws.cell(row=row, column=13).value
        s_cost = ws.cell(row=row, column=19).value
        
        a = actors.get(actor, {})
        
        # Determine correct K, L, M
        correct_l = determine_L(a)
        correct_m = determine_M(a)
        correct_k = determine_K(a)
        
        # Get weapon stats from yaml
        actor_weapons = a.get('weapons', [])
        primary_weapon = None
        for wname in actor_weapons:
            w = resolve_weapon(wname, all_weapons)
            if w and w.get('spread_damages'):
                primary_weapon = wname
                break
        
        if primary_weapon:
            w = resolve_weapon(primary_weapon, all_weapons)
            # Sheet Damage = sum of SpreadDamage × burst + DamagesConcrete
            total_spread = sum(w['spread_damages'])
            burst = w['burst']
            burst_delays = w['burst_delays']
            reload = w['reload']
            concrete = w['damages_concrete']
            
            # Sheet Damage = (sum SpreadDamage + DamagesConcrete) × burst
            yaml_damage = (total_spread + concrete) * burst
            # Sheet ReloadDelay = reload + (burst-1) × burst_delay
            if burst > 1 and burst_delays:
                yaml_reload = reload + (burst - 1) * burst_delays[0]
            else:
                yaml_reload = reload
            yaml_range = w['range']
        else:
            yaml_damage = 0
            yaml_reload = 0
            yaml_range = 0
        
        yaml_hp = a.get('hp')
        yaml_speed = a.get('speed')
        yaml_cost = a.get('cost')
        
        # Use yaml stats as the basis (they're the actual game values)
        hp = yaml_hp if yaml_hp else s_hp
        speed = yaml_speed if yaml_speed else s_speed
        damage = yaml_damage if yaml_damage else s_dmg
        reload = yaml_reload if yaml_reload else s_reload
        cost = s_cost  # Keep sheet target cost
        wclass = s_wclass  # Keep sheet weapon class
        k = correct_k
        l = correct_l
        m = correct_m
        
        # Solve for Range
        if damage > 0 and reload > 0:
            rng_solved = solve_range(hp, speed, damage, reload, wclass, k, l, m, cost)
            r_calc, dps = compute_cost(hp, speed, rng_solved, damage, reload, wclass, k, l, m)
        else:
            rng_solved = 0
            r_calc = 0
            dps = 0
        
        # Check for changes needed
        unit_changes = []
        if s_k != k:
            unit_changes.append(f"K: {s_k}->{k}")
        if s_l != l:
            unit_changes.append(f"L: {s_l}->{l}")
        if s_m != m:
            unit_changes.append(f"M: {s_m}->{m}")
        if s_dmg != damage:
            unit_changes.append(f"Dmg: {s_dmg}->{damage}")
        if s_reload != reload:
            unit_changes.append(f"Rld: {s_reload}->{reload}")
        if s_hp != hp:
            unit_changes.append(f"HP: {s_hp}->{hp}")
        if s_speed != speed:
            unit_changes.append(f"Spd: {s_speed}->{speed}")
        if s_cost != yaml_cost and yaml_cost:
            unit_changes.append(f"Cost: sheet={s_cost} yaml={yaml_cost}")
        
        if rng_solved and abs(rng_solved - s_rng) > 0.001:
            unit_changes.append(f"Rng: {s_rng:.3f}->{rng_solved:.3f}")
        
        # Check formula references
        o_formula = ws.cell(row=row, column=15).value
        if o_formula and isinstance(o_formula, str) and o_formula.startswith('='):
            if f"D{row}" not in o_formula:
                unit_changes.append(f"FORMULA BROKEN (refs wrong row)")
        
        status = "OK" if not unit_changes else "CHANGE"
        print(f"Row {row:2d} {actor:<35} {status}")
        if unit_changes:
            for c in unit_changes:
                print(f"       {c}")
            print(f"       RngSolved={rng_solved:.3f} (wdist={int(rng_solved*1000)}) R={r_calc:.1f} S={cost}")
        
        changes.append({
            'actor': actor, 'row': row,
            'hp': hp, 'speed': speed, 'damage': damage, 'reload': reload,
            'range': rng_solved, 'k': k, 'l': l, 'm': m, 'cost': cost,
            'wclass': wclass, 'yaml_cost': yaml_cost,
            'changes': unit_changes,
        })
    
    # Summary
    total_changes = sum(1 for c in changes if c['changes'])
    print(f"\n=== SUMMARY: {total_changes}/{len(changes)} units need changes ===")

if __name__ == "__main__":
    main()
