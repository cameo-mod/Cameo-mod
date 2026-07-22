#!/usr/bin/env python3
"""seed_design.py — Balance Pipeline Phase 3 (BALANCE_PIPELINE.md §6).

Seeds the ledger's design.* judgment fields (unit_class, special,
tech_tier) from the LEGACY workbook (docs/design/cameo_armor_system.xlsx)
and writes the discrepancy report docs/balance/discrepancies.md.

Matching:
- CABAL tab carries actor ids (column C) -> direct.
- Legacy type tabs (Infantry/Tanks/Vehicles/Aircraft/Defenses) carry
  display names only -> normalized-name match against ledger display
  names; ambiguous or unmatched rows are REPORTED, never guessed.
- docs/balance/name_map.yaml (committed, hand-curated) overrides:
      legacy display name: actor_id     # or "-" to ignore the row
Respects the ~$ Excel lock (queues per the balance law).
"""
from __future__ import annotations

import json
import pathlib
import re
import sys

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parents[2]
LEDGER = ROOT / "docs/balance"
LEGACY = ROOT / "docs/design/cameo_armor_system.xlsx"
NAME_MAP = LEDGER / "name_map.yaml"
REPORT = LEDGER / "discrepancies.md"

TYPE_TABS = {"Infantry": ("TD", ), "Tanks": (), "Vehicles": (), "Aircraft": (), "Defenses": ()}
FALLBACK_CATEGORY = {"infantry": "Infantry", "vehicles": "Vehicles", "aircraft": "Aircraft",
                     "naval": "Naval", "defenses": "Defenses"}
# legacy type-tab columns (recon 2026-07-18): B name, D hp, S cost,
# H weapon class, K special, L unit class, M tech tier
# CABAL tab: A mod, B name, C actor, ... K special, L unit class, M tier, R cost


def norm(s: str) -> str:
    s = re.sub(r"\(.*?\)", "", str(s)).lower()
    return re.sub(r"[^a-z0-9]", "", s)


def load_name_map() -> dict[str, str]:
    out = {}
    if NAME_MAP.exists():
        for line in NAME_MAP.read_text(encoding="utf-8").splitlines():
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            k, _, v = s.partition(":")
            out[norm(k)] = v.strip()
    return out


def ledger_docs() -> dict[str, dict]:
    out = {}
    for p in sorted(LEDGER.glob("*.json")):
        doc = json.loads(p.read_text(encoding="utf-8"))
        if "sections" in doc:  # skip registry files (class_anchors.json …)
            out[p.stem] = doc
    return out


def all_units(docs) -> dict[str, tuple[str, str, dict]]:
    """actor -> (ledger, section, unit)"""
    out = {}
    for name, doc in docs.items():
        for section, sec in doc["sections"].items():
            for actor, u in sec.items():
                out[actor] = (name, section, u)
    return out


def fnum(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main() -> int:
    if (LEGACY.parent / ("~$" + LEGACY.name)).exists():
        print("legacy workbook is OPEN in Excel (~$ lock) — queueing per the balance law; rerun later.")
        return 2
    docs = ledger_docs()
    units = all_units(docs)
    by_norm_name: dict[str, list[str]] = {}
    for actor, (_, _, u) in units.items():
        if u.get("name"):
            by_norm_name.setdefault(norm(u["name"]), []).append(actor)
    nmap = load_name_map()

    wb = openpyxl.load_workbook(LEGACY, data_only=False)
    seeded, direct, mapped, unmatched, ambiguous, dead = 0, 0, 0, [], [], []
    mismatches = []

    def seed(actor, special, unit_class, tier, legacy_cost, tab, rowname,
             category=None, subtype=None):
        nonlocal seeded
        if actor not in units:
            dead.append(f"{tab}: `{rowname}` -> `{actor}` (no such actor)")
            return
        _, _, u = units[actor]
        d = u.setdefault("design", {})
        for key, val in (("special", fnum(special)), ("unit_class", fnum(unit_class)),
                         ("tech_tier", fnum(tier)), ("category", category),
                         ("subtype", subtype)):
            if val is not None and d.get(key) is None:
                d[key] = val
        seeded += 1
        cost = fnum((u.get("cost") or {}).get("v"))
        lcost = fnum(legacy_cost)
        if cost is not None and lcost is not None and abs(cost - lcost) > 0.5:
            mismatches.append(f"`{actor}`: yaml cost {cost:.0f} vs legacy sheet {lcost:.0f} ({tab})")

    # --- CABAL tab: direct actor ids ---
    ws = wb["CABAL"]
    for r in range(2, ws.max_row + 1):
        actor = ws.cell(row=r, column=3).value
        if not actor:
            continue
        seed(str(actor).strip(), ws.cell(row=r, column=11).value,
             ws.cell(row=r, column=12).value, ws.cell(row=r, column=13).value,
             ws.cell(row=r, column=18).value, "CABAL", ws.cell(row=r, column=2).value)
        direct += 1

    # --- legacy type tabs: name matching ---
    for tab in ("Infantry", "Tanks", "Vehicles", "Aircraft", "Defenses"):
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        subtype = "Unclassified"
        for r in range(3, ws.max_row + 1):
            rowname = ws.cell(row=r, column=2).value
            hp = ws.cell(row=r, column=4).value
            if rowname and fnum(hp) is None:
                subtype = str(rowname).strip()
                continue
            if not rowname or fnum(hp) is None:
                continue
            key = norm(rowname)
            if key in nmap:
                target = nmap[key]
                if target == "-":
                    continue
                seed(target, ws.cell(row=r, column=11).value,
                     ws.cell(row=r, column=12).value, ws.cell(row=r, column=13).value,
                     ws.cell(row=r, column=19).value, tab, rowname, tab, subtype)
                mapped += 1
                continue
            cands = by_norm_name.get(key, [])
            if len(cands) == 1:
                seed(cands[0], ws.cell(row=r, column=11).value,
                     ws.cell(row=r, column=12).value, ws.cell(row=r, column=13).value,
                     ws.cell(row=r, column=19).value, tab, rowname, tab, subtype)
                mapped += 1
            elif len(cands) > 1:
                ambiguous.append(f"{tab}: `{rowname}` -> {cands}")
            else:
                unmatched.append(f"{tab}: `{rowname}`")

    for actor, (_, section, u) in units.items():
        category = FALLBACK_CATEGORY.get(section)
        if category:
            d = u.setdefault("design", {})
            if d.get("category") is None:
                d["category"] = category
            if d.get("subtype") is None:
                d["subtype"] = "Unclassified"

    # write ledgers back (design fields only changed)
    for name, doc in docs.items():
        p = LEDGER / f"{name}.json"
        p.write_text(json.dumps(doc, sort_keys=True, indent=1, ensure_ascii=False) + "\n",
                     encoding="utf-8", newline="\n")

    # unseeded units (in yaml, never in legacy sheet)
    unseeded = [a for a, (_, sec, u) in sorted(units.items())
                if sec in ("infantry", "vehicles", "aircraft", "naval", "defenses")
                and all(v is None for v in (u.get("design") or {}).values())]

    lines = ["# Balance ledger — legacy-sheet discrepancy report",
             "", f"_generated by seed_design.py; seeded {seeded} units "
             f"({direct} by actor id, {mapped} by name)_", "",
             "## Legacy rows with NO matching actor (dead or unmatched — extend name_map.yaml)", ""]
    lines += [f"- {x}" for x in unmatched + dead] or ["- none"]
    lines += ["", "## Ambiguous name matches (resolve in name_map.yaml)", ""]
    lines += [f"- {x}" for x in ambiguous] or ["- none"]
    lines += ["", "## Cost mismatches (yaml vs legacy sheet — maintainer picks the law)", ""]
    lines += [f"- {x}" for x in mismatches] or ["- none"]
    lines += ["", f"## Combat units never priced in the legacy sheet ({len(unseeded)})", ""]
    lines += [f"- `{a}`" for a in unseeded[:200]]
    if len(unseeded) > 200:
        lines.append(f"- … and {len(unseeded) - 200} more")
    REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")
    print(f"seeded {seeded} (direct {direct}, name-matched {mapped}); "
          f"unmatched {len(unmatched)}, ambiguous {len(ambiguous)}, dead {len(dead)}, "
          f"cost mismatches {len(mismatches)}, unseeded combat units {len(unseeded)}")
    print(f"report -> {REPORT.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
