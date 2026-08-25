#!/usr/bin/env python3
"""import_workbook.py — Balance Pipeline Phase 4a (BALANCE_PIPELINE.md §4).

A generated balance workbook -> ledger (docs/balance/*.json), input cells only.

Reads the designated INPUT cells (everything else is locked in the
generated sheet): unit HP / Speed / TechTier / UnitClass / Special /
Cost and weapon Damage / Reload / Burst / BurstDelays / Range /
WeaponClass. Prints every change; nothing else in the ledger moves.

Damage convention (BALANCE_PIPELINE §3): the sheet's Damage cell is the
per-shot TOTAL = SUM of the main offensive warheads (formula.spread_damage_sum).
When it changes, formula.distribute_damage() gives every main warhead the
same share total/N snapped to the current 100-damage grid. Folded
`PercentageScale` damage follows the main Damage automatically; standalone
percentage hits and extra-damage chips are not rewritten. The importer never
creates or fine-tunes an actor FirepowerMultiplier. A single number can never
again be broadcast identically onto every warhead (the 2026-07-22 over-damage
bug).
"""
from __future__ import annotations

import argparse
import json
import math
import pathlib
import sys

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "tools/balance"))
import formula  # noqa: E402
import tier_chain  # noqa: E402

LEDGER = ROOT / "docs/balance"
FACTION_WB = ROOT / "docs/design/cameo_balance_by_faction.xlsx"
TYPE_WB = ROOT / "docs/design/cameo_balance_by_type.xlsx"
TYPE_SHEETS = ("Infantry", "Tanks", "Vehicles", "Aircraft", "Defenses", "Naval")

# Never hard-code workbook column numbers here. The generated workbench has
# changed shape more than once, and stale indices turn a no-op import into a
# destructive stat shuffle. Resolve the contract from the actual header row.
IDENTITY_HEADERS = {"actor": "Actor", "slot": "Name"}
UNIT_HEADERS = {
    "hp": "HP", "speed": "Speed", "tech_tier": "TechTier",
    "unit_class": "UnitClass", "special": "Special", "cost": "Cost",
}
DESIGN_TEXT_HEADERS = {"subtype": "Class"}
WEAP_HEADERS = {
    "damage": "Damage", "reloaddelay": "Reload", "burst": "Burst",
    "burstdelays": "BurstDel", "range": "Range(wd)",
    "weapon_class": "WeapClass",
}


def field_columns(ws, fields: dict[str, str]) -> dict[str, int]:
    """Resolve logical importer fields from a generated sheet's header row."""
    headers = {}
    for column in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=column).value
        if value is not None:
            headers[str(value).strip()] = column
    missing = [header for header in fields.values() if header not in headers]
    if missing:
        raise ValueError(
            "workbook is missing required generated header(s): " + ", ".join(missing))
    return {field: headers[header] for field, header in fields.items()}


def fnum(v):
    try:
        f = float(v)
        if not math.isfinite(f):
            return None
        return int(f) if f == int(f) else f
    except (OverflowError, TypeError, ValueError):
        return None


def range_num(v):
    """Numeric WDist from workbook or OpenRA cell notation in the ledger."""
    return formula.wdist_value(v)


def burst_delays(v):
    """Normalize one workbook BurstDelays cell without losing a delay list."""
    return formula.burst_delays_text(v)


def _blank(value) -> bool:
    return value is None or str(value).strip() == ""


def _validate_sheet_inputs(ws, units, identity_cols, unit_cols, weap_cols) -> None:
    """Reject edits that could not be loaded by the engine before mutating data."""
    row = 2
    while row <= ws.max_row:
        actor_value = ws.cell(row=row, column=identity_cols["actor"]).value
        if not actor_value or str(actor_value).lstrip().startswith("↳"):
            row += 1
            continue
        actor = str(actor_value)
        unit = units.get(actor)
        if unit is None:
            row += 1
            continue

        for field, column in unit_cols.items():
            raw = ws.cell(row=row, column=column).value
            if _blank(raw):
                continue
            label = f"{actor}.{field}"
            if field in {"hp", "speed", "cost"}:
                formula.parse_int32(raw, label)
            elif fnum(raw) is None:
                raise ValueError(f"{label} must be a finite number")

        arms = {arm["slot"]: arm for arm in unit.get("armaments", [])}
        weapon_row = row + 1
        while (weapon_row <= ws.max_row and str(ws.cell(
                row=weapon_row, column=identity_cols["actor"]).value or
                "").lstrip().startswith("↳")):
            slot = str(ws.cell(
                row=weapon_row, column=identity_cols["slot"]).value)
            arm = arms.get(slot)
            if arm is not None:
                parsed = {}
                for field, column in weap_cols.items():
                    raw = ws.cell(row=weapon_row, column=column).value
                    if _blank(raw):
                        parsed[field] = None
                        continue
                    label = f"{actor}/{slot}.{field}"
                    if field in {"damage", "reloaddelay", "burst"}:
                        parsed[field] = formula.parse_int32(raw, label)
                    elif field == "burstdelays":
                        values = formula.burst_delay_values(raw)
                        if values is None:
                            raise ValueError(f"{label} must be an integer list")
                        parsed[field] = values
                    elif field == "range":
                        parsed[field] = formula.parse_wdist(raw)
                    else:
                        parsed[field] = fnum(raw)
                        if parsed[field] is None:
                            raise ValueError(f"{label} must be a finite number")

                reload_delay = parsed["reloaddelay"]
                if reload_delay is None:
                    reload_delay = formula.parse_int32(
                        arm.get("reloaddelay"), f"{actor}/{slot}.reloaddelay",
                        int(formula.ENGINE_DEFAULT_RELOAD_DELAY))
                burst = parsed["burst"]
                if burst is None:
                    burst = formula.parse_int32(
                        arm.get("burst"), f"{actor}/{slot}.burst",
                        formula.ENGINE_DEFAULT_BURST)
                delays = parsed["burstdelays"]
                if delays is None:
                    # Clearing the cell selects WeaponInfo's default delay.
                    delays = [int(formula.ENGINE_DEFAULT_BURST_DELAY)]

                if reload_delay <= 0:
                    raise ValueError(
                        f"{actor}/{slot}.reloaddelay must be greater than zero")
                if burst > 1 and len(delays) > 1 and len(delays) != burst - 1:
                    raise ValueError(
                        f"{actor}/{slot}.burstdelays must contain one value or "
                        f"Burst - 1 values ({burst - 1})")
            weapon_row += 1
        row = weapon_row


def displayed_tier_map(docs: list[dict]) -> dict[str, float]:
    """The TechTier values the generator displays before any workbook edit.

    A blank raw ``design.tech_tier`` falls back to a computed value in the
    derived sidecar. Writing that unchanged displayed value back as a manual
    override would freeze the derived result and make a no-op import mutate the
    ledger, so the importer must compare against the same precedence chain.
    """
    out = {}
    for doc in docs:
        dfile = LEDGER / "derived" / f"{doc['ledger']}.json"
        try:
            derived = json.loads(dfile.read_text(encoding="utf-8")) if dfile.is_file() else {}
        except (OSError, ValueError):
            derived = {}
        derived_sections = derived.get("sections") or {}
        for section, actors in doc.get("sections", {}).items():
            derived_actors = derived_sections.get(section) or {}
            for actor, unit in actors.items():
                design = unit.get("design") or {}
                derived_actor = derived_actors.get(actor) or {}
                out[actor] = tier_chain.effective_tier(
                    design.get("tech_tier"), derived_actor.get("tier_multiplier"),
                    default=1.0)
    return out


def import_sheet(ws, units, displayed_tiers: dict[str, float] | None = None) \
        -> tuple[bool, int, set[str]]:
    identity_cols = field_columns(ws, IDENTITY_HEADERS)
    unit_cols = field_columns(ws, UNIT_HEADERS)
    design_text_cols = field_columns(ws, DESIGN_TEXT_HEADERS)
    weap_cols = field_columns(ws, WEAP_HEADERS)
    _validate_sheet_inputs(ws, units, identity_cols, unit_cols, weap_cols)
    touched = False
    changes = 0
    changed_actors = set()
    r = 2
    while r <= ws.max_row:
        actor = ws.cell(row=r, column=identity_cols["actor"]).value
        if not actor or str(actor).lstrip().startswith("↳"):
            r += 1
            continue
        actor = str(actor)
        u = units.get(actor)
        if u is None:
            r += 1
            continue
        d = u.setdefault("design", {})
        for field, col in design_text_cols.items():
            v = ws.cell(row=r, column=col).value
            v = str(v).strip() if v is not None else ""
            displayed_old = d.get(field) or "Unclassified"
            if v and displayed_old != v:
                print(f"  {actor}.design.{field}: {displayed_old} -> {v}")
                d[field] = v
                touched = True
                changed_actors.add(actor)
        for field, col in unit_cols.items():
            raw_value = ws.cell(row=r, column=col).value
            if field in {"hp", "speed", "cost"} and not _blank(raw_value):
                v = formula.parse_int32(raw_value, f"{actor}.{field}")
            else:
                v = fnum(raw_value)
            if v is None:
                continue
            if field in ("tech_tier", "unit_class", "special"):
                # TechTier may be computed from the prerequisite chain when no
                # manual override exists. An unchanged generated cell is not an
                # edit and must not be written back as a permanent override.
                if field == "tech_tier":
                    baseline = (displayed_tiers or {}).get(actor)
                    if baseline is None:
                        baseline = fnum(d.get(field)) or 1
                    if abs(float(baseline) - float(v)) <= 1e-9:
                        continue
                    # A derived tier can differ from 1. Editing it to 1 is a real
                    # manual override, not the ordinary missing-value default.
                    should_write = fnum(d.get(field)) != v
                else:
                    baseline = fnum(d.get(field)) or 1
                    should_write = baseline != v
                if should_write:
                    print(f"  {actor}.design.{field}: {d.get(field)} -> {v}")
                    d[field] = v
                    touched = True
                    changed_actors.add(actor)
                continue
            slot = u.get(field) or u.get("speed_air") if field == "speed" else u.get(field)
            if field == "speed" and u.get("speed") is None and u.get("speed_air") is not None:
                slot = u["speed_air"]
            if slot is None:
                continue
            old = fnum(slot.get("v"))
            if old is not None and old != v:
                print(f"  {actor}.{field}: {old} -> {v}")
                slot["v"] = v
                touched = True
                changed_actors.add(actor)
                changes += 1
        wr = r + 1
        arms = {a["slot"]: a for a in u.get("armaments", [])}
        while (wr <= ws.max_row and
               str(ws.cell(row=wr, column=identity_cols["actor"]).value or
                   "").lstrip().startswith("↳")):
            slot_key = str(ws.cell(row=wr, column=identity_cols["slot"]).value)
            arm = arms.get(slot_key)
            if arm is not None:
                for field, col in weap_cols.items():
                    cell_value = ws.cell(row=wr, column=col).value
                    if field == "burstdelays":
                        if cell_value is None or str(cell_value).strip() == "":
                            # A blank generated cell means the engine default.
                            # When an explicit delay is cleared, persist that
                            # default as 5 so apply_balance can actually replace
                            # the old YAML value instead of losing the edit.
                            old_delay = burst_delays(arm.get(field))
                            default_delay = str(int(formula.ENGINE_DEFAULT_BURST_DELAY))
                            if old_delay is not None and old_delay != default_delay:
                                print(f"  {actor}/{slot_key}.{field}: "
                                      f"{old_delay} -> {default_delay} (engine default)")
                                arm[field] = default_delay
                                touched = True
                                changed_actors.add(actor)
                                changes += 1
                            continue
                        v = burst_delays(cell_value)
                        if v is None:
                            raise ValueError(
                                f"invalid BurstDel value for {actor}/{slot_key}: "
                                f"{cell_value!r}")
                    elif field == "range":
                        v = range_num(cell_value)
                    elif field in {"damage", "reloaddelay", "burst"}:
                        v = formula.parse_int32(
                            cell_value, f"{actor}/{slot_key}.{field}")
                    else:
                        v = fnum(cell_value)
                    if v is None:
                        continue
                    if field == "weapon_class":
                        if fnum(arm.get("design_weapon_class")) != v and not (
                                arm.get("design_weapon_class") is None and v == 1):
                            print(f"  {actor}/{slot_key}.weapon_class: "
                                  f"{arm.get('design_weapon_class')} -> {v}")
                            arm["design_weapon_class"] = v
                            touched = True
                            changed_actors.add(actor)
                        continue
                    if field == "damage":
                        # Sheet cell is the per-shot TOTAL (sum of main
                        # warheads). Split a new total back across the
                        # warheads via the ONE canonical splitter.
                        warheads = arm.get("damage_warheads", [])
                        old_total = formula.spread_damage_sum(warheads)
                        if old_total and abs(v - old_total) > 1e-9:
                            new = formula.distribute_damage(v, warheads)
                            print(f"  {actor}/{slot_key}.damage(total): "
                                  f"{int(old_total)} -> {int(v)} "
                                  f"(split across {len(new)} warhead(s))")
                            for w in arm["damage_warheads"]:
                                tag = w.get("tag")
                                if tag in new:
                                    w["damage"] = str(int(new[tag]))
                            touched = True
                            changed_actors.add(actor)
                            changes += 1
                        continue
                    if field == "burstdelays":
                        old = burst_delays(arm.get(field))
                        if old != v:
                            print(f"  {actor}/{slot_key}.{field}: {old} -> {v}")
                            arm[field] = v
                            touched = True
                            changed_actors.add(actor)
                            changes += 1
                        continue
                    old = (range_num(arm.get(field)) if field == "range"
                           else fnum(arm.get(field)))
                    # WeaponInfo.Burst defaults to one, and the generator shows
                    # that effective default instead of a blank cell. Compare
                    # against the same displayed baseline so a fresh workbook is
                    # a no-op, but an edit from the displayed 1 to (say) 2 can
                    # create the previously absent ledger/YAML field. Other
                    # numeric weapon cells display blank when absent; entering a
                    # value there is likewise an intentional field creation.
                    defaults = {
                        "reloaddelay": formula.ENGINE_DEFAULT_RELOAD_DELAY,
                        "burst": formula.ENGINE_DEFAULT_BURST,
                        "range": formula.ENGINE_DEFAULT_RANGE,
                    }
                    displayed_old = defaults.get(field) if old is None else old
                    if displayed_old != v:
                        print(f"  {actor}/{slot_key}.{field}: "
                              f"{displayed_old} -> {v}")
                        arm[field] = str(v)
                        touched = True
                        changed_actors.add(actor)
                        changes += 1
            wr += 1
        r = wr
    return touched, changes, changed_actors


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", choices=("faction", "type"), default="faction")
    args = ap.parse_args()
    workbook = TYPE_WB if args.workbook == "type" else FACTION_WB
    if not workbook.exists():
        print("no workbench found — run build_workbook.py first")
        return 2
    try:
        import openpyxl
    except ModuleNotFoundError:
        print("openpyxl not installed (pip install openpyxl)")
        return 2
    wb = openpyxl.load_workbook(workbook, data_only=False)
    docs = []
    units = {}
    owners = {}
    for jf in sorted(LEDGER.glob("*.json")):
        doc = json.loads(jf.read_text(encoding="utf-8"))
        if "sections" not in doc:
            continue
        docs.append((jf, doc))
        for sec in doc["sections"].values():
            units.update(sec)
            owners.update({actor: jf for actor in sec})
    tiers = displayed_tier_map([doc for _jf, doc in docs])
    total_changes = 0
    changed_docs = set()
    if args.workbook == "type":
        for sheet in TYPE_SHEETS:
            if sheet not in wb.sheetnames:
                continue
            touched, changes, changed_actors = import_sheet(wb[sheet], units, tiers)
            if touched:
                total_changes += changes
                changed_docs.update(owners[actor] for actor in changed_actors if actor in owners)
    else:
        for jf, doc in docs:
            name = doc["ledger"][:31]
            if name not in wb.sheetnames:
                continue
            faction_units = {}
            for sec in doc["sections"].values():
                faction_units.update(sec)
            touched, changes, _ = import_sheet(wb[name], faction_units, tiers)
            if touched:
                total_changes += changes
                changed_docs.add(jf)
    for jf, doc in docs:
        if jf in changed_docs:
            jf.write_text(json.dumps(doc, sort_keys=True, indent=1, ensure_ascii=False) + "\n",
                          encoding="utf-8", newline="\n")
            print(f"updated {jf.name}")
    print(f"import complete: {total_changes} stat changes")
    return 0


if __name__ == "__main__":
    sys.exit(main())
