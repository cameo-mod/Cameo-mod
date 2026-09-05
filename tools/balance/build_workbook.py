#!/usr/bin/env python3
"""build_workbook.py — Balance Pipeline Phase 2 (BALANCE_PIPELINE.md §3).

Ledger (docs/balance/*.json) -> docs/design/cameo_balance_by_faction.xlsx and docs/design/cameo_balance_by_type.xlsx.

The tracked files are generated WORKBENCHES, never authoritative sources:
regenerate them from the ledger whenever the model changes. Raw stats appear as editable cells; every derived
quantity is a live Excel formula (identical math to formula.py — proven
by test_formula.py's Tiger identity + symbolic equivalence). Layout:

  unit row:    Mod | Actor | Name | Class | HP | Speed | Armor | TechTier |
               UnitClass | Special | ... | O | P | Q | Price | Cost |
               Delta | Delta% | RangeSolver
  weapon rows: indented under the unit — Damage | Reload | Burst |
               BurstDelays | Range(wdist) | WeaponClass | EffReload* |
               DPS*        (* = formula cells)

Editable (unlocked) cells: unit HP/Speed/TechTier/UnitClass/Special,
weapon Damage/Reload/Burst/BurstDelays/Range/WeaponClass, and Cost.
Everything else is locked (sheet protection, no password).

Damage convention: the weapon row's Damage cell is the SUM of the main
offensive warheads. Importing a changed total redistributes it as equal
100-grid shares across those main warheads (Phase 4); percentage and
extra-damage applications keep their dedicated rules.

Defense rows with no Speed get the legacy speed=100 convention as an
explicit input value (until the Formula-v2 defense class replaces it).
"""
from __future__ import annotations

import argparse
import hashlib
import json
import pathlib
import re
import sys

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
import tier_chain
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "tools/balance"))
import formula  # noqa: E402

LEDGER = ROOT / "docs/balance"
MOD = ROOT / "mods/cameo"
MOD_CONFIG = MOD / "mod.yaml"
DEFAULTS = MOD / "rules/defaults.yaml"
OUTFILE = ROOT / "docs/design/cameo_balance_by_faction.xlsx"
TYPE_OUTFILE = ROOT / "docs/design/cameo_balance_by_type.xlsx"

SECTION_ORDER = ("infantry", "vehicles", "aircraft", "naval", "defenses",
                 "buildings", "upgrades", "promotions", "misc", "faction")
TYPE_ORDER = ("Infantry", "Tanks", "Vehicles", "Aircraft", "Defenses", "Naval")

HDR = ["Mod", "Actor", "Name", "Class", "HP", "Speed", "Armor",
       "TechTier", "UnitClass", "Special", "FirepowerMultiplier",
       "Damage", "Reload", "Burst", "BurstDel",
       "Range(wd)", "WeapClass", "EffReload", "DPS",
       "O", "P", "Q", "Price", "Cost", "Delta", "Delta%", "RangeSolve",
       "WeaponTypes"]
COL = {name: i + 1 for i, name in enumerate(HDR)}
UNLOCKED_UNIT = ("Class", "HP", "Speed", "TechTier", "UnitClass", "Special",
                 "Cost")
UNLOCKED_WEAPON = ("Damage", "Reload", "Burst", "BurstDel", "Range(wd)", "WeapClass")

DEFAULT_FONT = Font(size=10)
HEAD_FONT = Font(bold=True, size=10)
SECTION_FILL = PatternFill("solid", fgColor="DDDDDD")
WEAPON_FONT = Font(italic=True, size=9)

COLUMN_WIDTHS = {
    "Mod": 10,
    "Actor": 30,
    "Name": 28,
    "Class": 22,
    "HP": 10,
    "Speed": 10,
    "Armor": 14,
    "TechTier": 12,
    "UnitClass": 12,
    "Special": 10,
    "FirepowerMultiplier": 18,
    "Damage": 12,
    "Reload": 12,
    "Burst": 10,
    "BurstDel": 12,
    "Range(wd)": 12,
    "WeapClass": 12,
    "EffReload": 12,
    "DPS": 12,
    "O": 12,
    "P": 12,
    "Q": 12,
    "Price": 12,
    "Cost": 12,
    "Delta": 12,
    "Delta%": 12,
    "RangeSolve": 14,
    "WeaponTypes": 38,
}


def L(name: str) -> str:
    return get_column_letter(COL[name])


def fnum(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def range_num(v):
    """Workbook-ready WDist, including OpenRA cell notation such as 40c0."""
    return formula.wdist_value(v)


def burst_delays_value(v):
    """Workbook input value without discarding a comma-separated delay list."""
    values = formula.burst_delay_values(v)
    if values is None:
        return None
    return values[0] if len(values) == 1 else formula.burst_delays_text(values)


def eff_reload_formula(row: int) -> str:
    """Excel form of formula.eff_reload, including lists and engine defaults."""
    reload_cell = f"{L('Reload')}{row}"
    burst_cell = f"{L('Burst')}{row}"
    delay_cell = f"{L('BurstDel')}{row}"
    gaps = f"MAX({burst_cell}-1,0)"
    # A scalar delay repeats for every gap. A text list contains one delay per
    # gap and is summed. Blank means WeaponInfo's default delay of five ticks.
    delay_total = (
        f"IF(ISBLANK({delay_cell}),{formula.ENGINE_DEFAULT_BURST_DELAY:g}*{gaps},"
        f"IF(ISNUMBER({delay_cell}),{delay_cell}*{gaps},"
        f"SUM(IFERROR(VALUE(TEXTSPLIT({delay_cell},\",\")),0))))")
    return f"={reload_cell}+IF({gaps}>0,{delay_total},0)"


def unit_rows(ws, theme, aid, u, section, row):
    """Write one unit (+ its weapon rows); returns next free row."""
    first = row
    ws.cell(row=row, column=COL["Mod"], value=theme)
    ws.cell(row=row, column=COL["Actor"], value=aid)
    ws.cell(row=row, column=COL["Name"], value=u.get("name") or aid)
    d = u.get("design") or {}
    ws.cell(row=row, column=COL["Class"], value=d.get("subtype") or "Unclassified")
    hp = fnum((u.get("hp") or {}).get("v"))
    speed = fnum((u.get("speed") or {}).get("v") or (u.get("speed_air") or {}).get("v"))
    if speed is None and section == "defenses":
        speed = 100.0  # legacy convention until the Formula-v2 defense class
    armor = (u.get("armor") or {}).get("v")
    cost = fnum((u.get("cost") or {}).get("v"))
    ws.cell(row=row, column=COL["HP"], value=hp)
    ws.cell(row=row, column=COL["Speed"], value=speed)
    ws.cell(row=row, column=COL["Armor"], value=armor)
    tech_tier = getattr(build, "_tier_map", {}).get(aid,
                                                    fnum(d.get("tech_tier")) or 1.0)
    ws.cell(row=row, column=COL["TechTier"], value=tech_tier)
    ws.cell(row=row, column=COL["UnitClass"], value=d.get("unit_class") or 1)
    ws.cell(row=row, column=COL["Special"], value=d.get("special") or 1)
    fp_raw = fnum((u.get("firepower_multiplier") or {}).get("v"))
    fp_factor = fp_raw / 100 if fp_raw is not None else None
    ws.cell(row=row, column=COL["FirepowerMultiplier"], value=fp_factor)
    ws.cell(row=row, column=COL["Cost"], value=cost)

    wrows = []
    wrows_all = []
    for arm in u.get("armaments", []):
        if arm.get("unresolved"):
            continue
        row += 1
        wrows_all.append(row)
        if arm.get("pricing", True):
            wrows.append(row)
        ws.cell(row=row, column=COL["Actor"],
                value=f"  ↳ {arm.get('weapon')}").font = WEAPON_FONT
        ws.cell(row=row, column=COL["Name"], value=arm.get("slot")).font = WEAPON_FONT
        # what the weapon DOES: its resolved ^-class templates (armor
        # profiles + effects) — e.g. "^SmallArms, ^Chaingun, ^LaserWeapon".
        wtypes = arm.get("warheads") or arm.get("versus_templates") or []
        wt = ws.cell(row=row, column=COL["WeaponTypes"],
                     value=", ".join(wtypes))
        wt.font = WEAPON_FONT
        if arm.get("requires"):
            wt.comment = Comment("fires when: " + str(arm.get("requires")),
                                 "balance-pipeline")
        warheads = arm.get("damage_warheads", [])
        damages = [fnum(w.get("damage")) for w in warheads]
        damages = [x for x in damages if x is not None]
        # Damage cell = per-shot TOTAL = SUM of the main offensive warheads
        # (formula.spread_damage_sum) — the SAME quantity pricing uses, so the
        # DPS cell below is consistent with the price. A weapon with only
        # side warheads falls back to max. See BALANCE_PIPELINE.md §3.
        total = formula.spread_damage_sum(warheads)
        n_main = sum(1 for w in warheads if formula._is_main_spread(w))
        cdmg = ws.cell(row=row, column=COL["Damage"],
                       value=int(total) if total else (max(damages) if damages else None))
        if len(damages) > 1:
            cdmg.comment = Comment(
                f"per-shot TOTAL across {n_main} main warhead(s).\n"
                "raw warhead damages: " + ", ".join(str(int(x)) for x in damages)
                + "\nediting this gives every main warhead the IDENTICAL value "
                "total/N snapped to the 100 grid (FriendlyFire + ExtraDamage "
                "twins 50%; percentage twins follow their node denominator; "
                "ExtraDamage excluded from the total). Actor FirepowerMultiplier "
                "is read for legacy compatibility, not used as a tuning knob.",
                "balance-pipeline")
        reload_delay = fnum(arm.get("reloaddelay"))
        burst = fnum(arm.get("burst"))
        weapon_range = range_num(arm.get("range"))
        ws.cell(
            row=row, column=COL["Reload"],
            value=(formula.ENGINE_DEFAULT_RELOAD_DELAY
                   if reload_delay is None else reload_delay))
        ws.cell(
            row=row, column=COL["Burst"],
            value=formula.ENGINE_DEFAULT_BURST if burst is None else burst)
        ws.cell(row=row, column=COL["BurstDel"],
                value=burst_delays_value(arm.get("burstdelays")))
        ws.cell(
            row=row, column=COL["Range(wd)"],
            value=formula.ENGINE_DEFAULT_RANGE if weapon_range is None else weapon_range)
        ws.cell(row=row, column=COL["WeapClass"],
                value=fnum(arm.get("design_weapon_class")) or 1)
        r = row
        fp_factor = f"IF(ISBLANK({L('FirepowerMultiplier')}{first}),1,{L('FirepowerMultiplier')}{first})"
        ws.cell(row=r, column=COL["EffReload"],
                value=eff_reload_formula(r))
        # No *WeapClass (W4): formula.dps() dropped it, and this sheet must emit
        # the SAME math as the module or the two silently disagree. The WeapClass
        # column stays as design data — it just no longer prices.
        ws.cell(row=r, column=COL["DPS"],
                value=f"=IFERROR({L('Damage')}{r}*MAX({L('Burst')}{r},1)"
                      f"/{L('EffReload')}{r}*{fp_factor},0)")

    if wrows and hp is not None and speed is not None:
        r = first
        dps_sum = "+".join(f"{L('DPS')}{w}" for w in wrows)
        rng = "MAX(" + ",".join(f"{L('Range(wd)')}{w}" for w in wrows) + ")"
        ws.cell(row=r, column=COL["DPS"], value=f"={dps_sum}")
        ws.cell(row=r, column=COL["Range(wd)"], value=f"={rng}")
        H_, S_, T_, U_, K_ = (f"{L('HP')}{r}", f"{L('Speed')}{r}", f"{L('TechTier')}{r}",
                              f"{L('UnitClass')}{r}", f"{L('Special')}{r}")
        RNG = f"({L('Range(wd)')}{r}/1000)"
        DPS_ = f"{L('DPS')}{r}"
        # Global O/P/Q remain as a reference; Price/RangeSolve use the
        # class-baseline formula when the class has a live spec anchor.
        ws.cell(row=r, column=COL["O"],
                value=f"=({H_}/100000+{S_}/100+{RNG}*{K_}/5+{DPS_}/200)*200*{U_}*{T_}")
        ws.cell(row=r, column=COL["P"],
                value=f"=(({H_}*{S_}/25000)+({RNG}*{K_}*{DPS_}/2.5))*{U_}*{T_}")
        ws.cell(row=r, column=COL["Q"],
                value=f"=({H_}*{S_}*{RNG}*{K_}*{DPS_}*{U_}*{T_})/12500000")
        anchors = getattr(build, "_anchors", None) or {}
        design = u.get("design") or {}
        cls = design.get("class_anchor") or subtype_to_anchor(design.get("subtype"))
        a = anchors.get(cls) if cls else None
        if a and a.get("spec"):
            # class-baseline price (formula.py class_baseline_price) is
            # linear in range, so the range solver is exact and rounded.
            s = a["spec"]
            hp0, sp0, dps0, rng0, c0 = (
                s["hp0"], s["speed0"], s["dps0"], s["range0_wdist"], s["cost0"])
            h = f"({H_}/{hp0})"
            sd = f"({S_}/{sp0})"
            d = f"({DPS_}/{dps0})"
            rn = f"(({L('Range(wd)')}{r}/{rng0})*{K_})"
            anchor_tier = getattr(build, "_anchor_tier_map", {}).get(cls, 1.0)
            # TechTier column is absolute f(C); class-baseline needs relative
            # f(C)/f(C_anchor), so divide by the anchor's absolute tier here.
            T_REL = f"({T_}/{anchor_tier})"
            # Some intentionally unarmed class anchors have zero DPS/range
            # baselines. Keep their price blank instead of emitting a visible
            # #DIV/0! until that class receives a combat-capable anchor.
            price = (f"=IFERROR((({h}+{sd}+{d}+{rn})*{c0}/4"
                     f"+(({h}*{sd})+({rn}*{d}))*{c0}/2"
                     f"+({h}*{sd}*{rn}*{d})*{c0})*{T_REL}/3,\"\")")
            A = f"(({h}+{sd}+{d})*{c0}/4+({h}*{sd})*{c0}/2)*{T_REL}"
            B = f"({c0}/4+({d})*{c0}/2+({h}*{sd}*{d})*{c0})*{T_REL}"
            rs = (f"=IFERROR(ROUND(((3*{L('Cost')}{r}-{A})/{B}*{rng0})"
                   f"/{K_},-1),\"\")")
            ws.cell(row=r, column=COL["Price"], value=price)
            ws.cell(row=r, column=COL["RangeSolve"], value=rs)
        else:
            # legacy: either the global Tiger formula or the old class_anchor
            # deviation form (used only for the mbt reference anchor).
            if a and a.get("signed_off") and a.get("cost0"):
                ws.cell(row=r, column=COL["Price"],
                        value=f"={a['cost0']}*({L('O')}{r}/{a['o0']}"
                              f"+{L('P')}{r}/{a['p0']}+{L('Q')}{r}/{a['q0']})/3")
            else:
                ws.cell(row=r, column=COL["Price"],
                        value=f"=({L('O')}{r}+{L('P')}{r}+{L('Q')}{r})/3")
            # closed-form range solver for the global price formula
            A = (f"((({H_}/100000+{S_}/100+{DPS_}/200)*200*{U_}*{T_})"
                 f"+(({H_}*{S_}/25000)*{U_}*{T_}))/3")
            B = (f"((({K_}/5)*200*{U_}*{T_})+(({K_}*{DPS_}/2.5)*{U_}*{T_})"
                 f"+(({H_}*{S_}*{K_}*{DPS_}*{U_}*{T_})/12500000))/3")
            ws.cell(row=r, column=COL["RangeSolve"],
                    value=f"=IFERROR(({L('Cost')}{r}-{A})/{B}*1000,\"\")")
        ws.cell(row=r, column=COL["Delta"],
                value=f"=IFERROR({L('Price')}{r}-{L('Cost')}{r},\"\")")
        ws.cell(row=r, column=COL["Delta%"],
                value=f"=IFERROR(ABS({L('Delta')}{r})/MAX({L('Cost')}{r},1),\"\")")
    return row + 1


def protect(ws, unit_cells, weapon_cells):
    for c in unit_cells:
        c.protection = Protection(locked=False)
    for c in weapon_cells:
        c.protection = Protection(locked=False)
    ws.protection.sheet = True
    ws.protection.enable()


def workbook_fingerprint() -> str:
    """Hash every source/input that can change generated workbook semantics."""
    paths = [pathlib.Path(__file__), pathlib.Path(formula.__file__),
             pathlib.Path(tier_chain.__file__), MOD_CONFIG, DEFAULTS]
    paths.extend(sorted(LEDGER.rglob("*.json")))
    digest = hashlib.sha256()
    for path in paths:
        digest.update(path.relative_to(ROOT).as_posix().encode("utf-8"))
        digest.update(b"\0")
        # Git may materialize these text files with LF or CRLF. Hash normalized
        # content so an unchanged tree produces the same workbook on every OS.
        normalized = path.read_text(encoding="utf-8").replace("\r\n", "\n").replace("\r", "\n")
        digest.update(normalized.encode("utf-8"))
        digest.update(b"\0")
    return digest.hexdigest()


def add_constants_sheet(wb, title):
    const = wb.active
    const.title = "Constants"
    notes = [
        (f"{title} — generated workbench (build_workbook.py).", None),
        ("TRACKED generated workbench; regenerate from the ledger, never treat as source.", None),
        ("Formula law (formula.py, Tiger anchor O=P=Q=Cost=800):", None),
        ("DPS", "Damage*Burst/(Reload+sum of every burst gap; blank delay = 5 each)*FirepowerMultiplier"),
        ("WeapClass", "design data only — retired from pricing 2026-08-11 (W4)"),
        ("O", "(HP/1e5+Speed/100+Rng*Spec/5+DPS/200)*200*UC*Tier"),
        ("P", "((HP*Speed/25000)+(Rng*Spec*DPS/2.5))*UC*Tier"),
        ("Q", "(HP*Speed*Rng*Spec*DPS*UC*Tier)/12.5e6"),
        ("Price", "(O+P+Q)/3   |   RangeSolve: price(range)=Cost, exact"),
        ("Generator fingerprint", workbook_fingerprint()),
    ]
    for i, (a, b) in enumerate(notes, start=1):
        const.cell(row=i, column=1, value=a).font = HEAD_FONT if i <= 3 else Font()
        if b:
            const.cell(row=i, column=2, value=b)


def setup_sheet(wb, title):
    ws = wb.create_sheet(title=title[:31])
    for c, h in enumerate(HDR, start=1):
        ws.cell(row=1, column=c, value=h).font = HEAD_FONT
    # Freeze the first three columns (Mod, Actor, Name) and the header row.
    ws.freeze_panes = "D2"
    return ws


def finish_sheet(ws, row, unit_unlock, weap_unlock):
    ws.conditional_formatting.add(
        f"{L('Delta%')}2:{L('Delta%')}{row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                       mid_type="num", mid_value=0.25, mid_color="FFEB84",
                       end_type="num", end_value=0.6, end_color="F8696B"))
    # Default font for the whole sheet, then explicit per-column widths
    # so long values (Superheavy armor, FirepowerMultiplier header, etc.)
    # are readable without being cut off.
    ws.sheet_format.defaultRowHeight = 15
    for col_name, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[L(col_name)].width = width
    ws.row_dimensions[1].height = 30
    for c in range(1, len(HDR) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    protect(ws, unit_unlock, weap_unlock)


def write_unit(ws, theme, aid, u, section, row, unit_unlock, weap_unlock):
    first = row
    row = unit_rows(ws, theme, aid, u, section, row)
    for col in UNLOCKED_UNIT:
        # HP/Speed/Cost need an existing provenance slot for apply_balance.py
        # to write safely. In particular, the displayed Speed 100 for a defense
        # without a movement trait is only a pricing convention, not an engine
        # field. Keep unsupported cells locked instead of accepting an edit the
        # importer would have to ignore.
        if col == "HP" and not isinstance(u.get("hp"), dict):
            continue
        if col == "Speed" and not (
                isinstance(u.get("speed"), dict) or
                isinstance(u.get("speed_air"), dict)):
            continue
        if col == "Cost" and not isinstance(u.get("cost"), dict):
            continue
        unit_unlock.append(ws.cell(row=first, column=COL[col]))
    resolved_arms = [arm for arm in u.get("armaments", []) if not arm.get("unresolved")]
    for offset, arm in enumerate(resolved_arms, 1):
        wr = first + offset
        for col in UNLOCKED_WEAPON:
            # There is no scalar Damage field to create when a weapon has no
            # existing main damage warhead. Other top-level weapon fields can be
            # inserted safely by apply_balance.py.
            if col == "Damage" and not formula.spread_damage_sum(
                    arm.get("damage_warheads", [])):
                continue
            weap_unlock.append(ws.cell(row=wr, column=COL[col]))
    return row


def load_faction_order():
    order = {}
    for line in MOD_CONFIG.read_text(encoding="utf-8-sig").splitlines():
        match = re.match(r"^Include:\s*(ContentPacks/.+)/content\.yaml\s*$", line)
        if match:
            pack = match.group(1).replace("\\", "/").casefold()
            order.setdefault(pack, len(order))
    return order


def template_category(name):
    if name in {"Medic", "Mechanic", "Dog"} or name.endswith("Infantry"):
        return "Infantry"
    if name.endswith("Tank") or name == "LineBreaker":
        return "Tanks"
    if name.endswith("Vehicle") or name in {"Harvester", "FireSupport", "Artillery"}:
        return "Vehicles"
    if name.endswith("Ship"):
        return "Naval"
    if name.endswith("Defense") or name == "Bunker":
        return "Defenses"
    if name.endswith(("Fighter", "Bomber", "Helicopter", "Spaceship", "AirUnit")) or \
            name == "FlyingInfantry":
        return "Aircraft"
    return None


def normalize_type_name(value):
    value = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", value)
    value = re.sub(r"[^a-z0-9]+", "", value.casefold())
    # Strip Template suffix before normalising so auto-derived subtypes
    # like "MainBattleTank" sort against the defaults.yaml entries.
    if value.endswith("template"):
        value = value[:-8]
    for suffix in ("infantry", "vehicles", "vehicle", "tanks", "tank", "defenses", "defense",
                   "ships", "ship", "aircraft", "units", "unit"):
        if value.endswith(suffix):
            value = value[:-len(suffix)]
            break
    if value.endswith("ies"):
        value = value[:-3] + "y"
    elif value.endswith("s"):
        value = value[:-1]
    return value


def subtype_to_anchor(st: str | None) -> str | None:
    """Map a design subtype to the class_anchor key when it is not explicit."""
    if not st:
        return None
    name = re.sub(r"[^A-Za-z0-9]", "", str(st)).casefold()
    exact = {
        "scoutinfantry": "scout",
        "closecombatinfantry": "closecombat",
        "specialforcesinfantry": "special_forces",
        "mainbattletank": "mbt",
        "linebreaker": "mbt",
    }
    return exact.get(name)


def load_template_order():
    order = {category: [] for category in TYPE_ORDER}
    for line in DEFAULTS.read_text(encoding="utf-8-sig").splitlines():
        match = re.match(r"^\^([A-Za-z0-9_]+)Template:", line)
        if not match:
            continue
        name = match.group(1)
        category = template_category(name)
        if category:
            order[category].append(normalize_type_name(name))
    return order


def subtype_sort_key(category, subtype, template_order):
    normalized = normalize_type_name(subtype)
    if normalized == "unclassified":
        return (2, 0, "")
    for index, template in enumerate(template_order[category]):
        if normalized == template or normalized in template or template in normalized:
            return (0, index, normalized)
    return (1, len(template_order[category]), normalized)


def faction_order_key(pack):
    normalized = pack.replace("\\", "/").casefold()
    marker = "contentpacks/"
    index = normalized.find(marker)
    return normalized[index:] if index >= 0 else normalized


def load_ledgers():
    faction_order = load_faction_order()
    docs = []
    for jf in LEDGER.glob("*.json"):
        doc = json.loads(jf.read_text(encoding="utf-8"))
        if "sections" in doc:
            docs.append(doc)
    return sorted(docs, key=lambda doc: (faction_order.get(faction_order_key(doc["pack"]), len(faction_order)),
                                         doc["ledger"]))


def load_tier_map(docs):
    """{(actor): absolute tier multiplier} from raw design + derived sidecar."""
    tier_map = {}
    for doc in docs:
        dfile = LEDGER / "derived" / f"{doc['ledger']}.json"
        try:
            ddoc = json.loads(dfile.read_text(encoding="utf-8")) if dfile.is_file() else {}
        except Exception:
            ddoc = {}
        dsec = ddoc.get("sections") or {}
        for section, sec in doc["sections"].items():
            du_sec = dsec.get(section) or {}
            for actor, u in sec.items():
                design = u.get("design") or {}
                du = du_sec.get(actor) or {}
                tier_map[actor] = tier_chain.effective_tier(
                    design.get("tech_tier"), du.get("tier_multiplier"), default=1.0)
    return tier_map


def load_anchor_tiers(anchors, tier_map):
    """{class: absolute anchor tier multiplier}."""
    out = {}
    for cls, a in anchors.items():
        anchor_actor = a.get("anchor_actor")
        if anchor_actor and anchor_actor in tier_map:
            out[cls] = tier_map[anchor_actor]
        else:
            out[cls] = fnum(a.get("tech_tier")) or 1.0
    return out


def write_faction_workbook(docs):
    wb = openpyxl.Workbook()
    add_constants_sheet(wb, "cameo_balance_by_faction")
    for doc in docs:
        ws = setup_sheet(wb, doc["ledger"])
        theme = doc["ledger"].split("_")[0]
        row = 2
        unit_unlock, weap_unlock = [], []
        for section in SECTION_ORDER:
            sec = doc["sections"].get(section)
            if not sec:
                continue
            ws.cell(row=row, column=1, value=section.upper()).font = HEAD_FONT
            for c in range(1, len(HDR) + 1):
                ws.cell(row=row, column=c).fill = SECTION_FILL
            row += 1
            for aid in sorted(sec):
                row = write_unit(ws, theme, aid, sec[aid], section, row, unit_unlock, weap_unlock)
        finish_sheet(ws, row, unit_unlock, weap_unlock)
    wb.save(OUTFILE)
    print(f"wrote {OUTFILE.relative_to(ROOT)} ({len(wb.sheetnames) - 1} faction tabs)")


def type_units(docs):
    grouped = {category: {} for category in TYPE_ORDER}
    for faction_index, doc in enumerate(docs):
        theme = doc["ledger"].split("_")[0]
        for section, sec in doc["sections"].items():
            for aid, u in sec.items():
                d = u.get("design") or {}
                subtype = d.get("subtype") or "Unclassified"
                category = template_category(subtype)
                if category not in grouped:
                    continue
                grouped[category].setdefault(subtype, []).append(
                    (faction_index, theme, aid, u, section))
    return grouped


def write_type_workbook(docs):
    wb = openpyxl.Workbook()
    add_constants_sheet(wb, "cameo_balance_by_type")
    template_order = load_template_order()
    for category, subtypes in type_units(docs).items():
        ws = setup_sheet(wb, category)
        row = 2
        unit_unlock, weap_unlock = [], []
        for subtype in sorted(subtypes, key=lambda name: subtype_sort_key(category, name, template_order)):
            ws.cell(row=row, column=1, value=subtype).font = HEAD_FONT
            for c in range(1, len(HDR) + 1):
                ws.cell(row=row, column=c).fill = SECTION_FILL
            row += 1
            for _, theme, aid, u, section in sorted(subtypes[subtype], key=lambda x: (x[0], x[2])):
                row = write_unit(ws, theme, aid, u, section, row, unit_unlock, weap_unlock)
        finish_sheet(ws, row, unit_unlock, weap_unlock)
    wb.save(TYPE_OUTFILE)
    print(f"wrote {TYPE_OUTFILE.relative_to(ROOT)} ({len(wb.sheetnames) - 1} type tabs)")


def expected_type_layout(docs):
    template_order = load_template_order()
    layout = {}
    for category, subtypes in type_units(docs).items():
        layout[category] = []
        for subtype in sorted(subtypes, key=lambda name: subtype_sort_key(category, name, template_order)):
            actors = [aid for _, _, aid, _, _ in sorted(subtypes[subtype], key=lambda x: (x[0], x[2]))]
            layout[category].append((subtype, actors))
    return layout


def actual_type_layout(ws):
    # Some standards-compliant XLSX writers omit the optional worksheet
    # ``dimension`` hint. Stream the read-only rows once instead of relying on
    # max_row (or repeatedly reparsing the XML through ws.cell()).
    layout = []
    subtype = None
    actors = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        left = values[COL["Mod"] - 1] if len(values) >= COL["Mod"] else None
        actor = values[COL["Actor"] - 1] if len(values) >= COL["Actor"] else None
        if left and not actor:
            if subtype is not None:
                layout.append((subtype, actors))
            subtype, actors = str(left), []
        elif subtype is not None and actor and not str(actor).lstrip().startswith("↳"):
            actors.append(str(actor))
    if subtype is not None:
        layout.append((subtype, actors))
    return layout


def check_order(docs):
    errors = []
    expected_fingerprint = workbook_fingerprint()
    if not OUTFILE.exists():
        errors.append(f"missing {OUTFILE.relative_to(ROOT)}")
    else:
        wb = openpyxl.load_workbook(OUTFILE, read_only=True, data_only=False)
        expected_tabs = ["Constants", *[doc["ledger"][:31] for doc in docs]]
        if wb.sheetnames != expected_tabs:
            errors.append("faction workbook tab order differs from mod.yaml load order")
        if wb["Constants"]["B10"].value != expected_fingerprint:
            errors.append("faction workbook was not regenerated from current code/ledger")
    if not TYPE_OUTFILE.exists():
        errors.append(f"missing {TYPE_OUTFILE.relative_to(ROOT)}")
    else:
        wb = openpyxl.load_workbook(TYPE_OUTFILE, read_only=True, data_only=False)
        expected_layout = expected_type_layout(docs)
        expected_tabs = ["Constants", *TYPE_ORDER]
        if wb.sheetnames != expected_tabs:
            errors.append("type workbook tab order differs from the declared major type order")
        if wb["Constants"]["B10"].value != expected_fingerprint:
            errors.append("type workbook was not regenerated from current code/ledger")
        for category, expected in expected_layout.items():
            if category not in wb.sheetnames:
                continue
            if actual_type_layout(wb[category]) != expected:
                errors.append(f"{category} workbook order differs from defaults.yaml templates, mod.yaml, or ledger")
    for error in errors:
        print(f"ORDER DRIFT: {error}")
    if errors:
        print("workbook order check: stale — run build_workbook.py")
        return 1
    print("workbook order/fingerprint check: current")
    return 0


def build(docs=None):
    anchors_file = LEDGER / "class_anchors.json"
    build._anchors = {k: v for k, v in (json.loads(
        anchors_file.read_text(encoding="utf-8")) if anchors_file.exists()
        else {}).items() if isinstance(v, dict)}
    docs = docs if docs is not None else load_ledgers()
    build._tier_map = load_tier_map(docs)
    build._anchor_tier_map = load_anchor_tiers(build._anchors, build._tier_map)
    OUTFILE.parent.mkdir(parents=True, exist_ok=True)
    write_faction_workbook(docs)
    write_type_workbook(docs)
    return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true", help="verify workbook ordering against YAML and ledger")
    args = ap.parse_args()
    docs = load_ledgers()
    return check_order(docs) if args.check else build(docs)


if __name__ == "__main__":
    sys.exit(main())
