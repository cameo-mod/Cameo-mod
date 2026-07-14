#!/usr/bin/env python3
"""
Compute O, P, Q, R from sheet inputs and compare R to S (target cost).
Also fix broken formula references in rows 27-29.
"""
import openpyxl
import os

def compute_cost(hp, speed, rng, damage, reload, wclass, k, l, m):
    """Compute O, P, Q, R from inputs."""
    dps = damage / reload * wclass
    
    o = (hp/100000 + speed/100 + rng*k/5 + dps/200) * 200 * l * m
    p = (hp*speed/25000 + rng*k*dps/2.5) * l * m
    q = hp*speed*rng*k*dps*l*m / 12500000
    r = (o + p + q) / 3
    
    return o, p, q, r, dps

def solve_range(hp, speed, damage, reload, wclass, k, l, m, cost):
    """Solve for Range from the cost identity."""
    dps = damage / reload * wclass
    h = hp / 100000
    s = speed / 100
    d = dps / 200
    p_const = hp * speed / 25000
    q_coeff = hp * speed * dps / 12500000
    
    # 3*Cost/(L*M) = 200*(h+s+d) + p_const + Range*K*(40 + DPS/2.5 + q_coeff)
    lhs = 3 * cost / (l * m)
    constant = 200 * (h + s + d) + p_const
    range_coeff = k * (40 + dps / 2.5 + q_coeff)
    
    if range_coeff == 0:
        return None
    
    range_val = (lhs - constant) / range_coeff
    return range_val

def main():
    wb = openpyxl.load_workbook('docs/design/cameo_armor_system.xlsx', data_only=False)
    ws = wb['CABAL']
    
    print("=== CABAL Cost Analysis: R vs S ===")
    print(f"{'Actor':<35} {'HP':>8} {'Spd':>5} {'Rng':>8} {'Dmg':>8} {'Rld':>5} {'K':>4} {'L':>5} {'M':>4} {'Cost(S)':>8} {'R(calc)':>10} {'Diff':>8} {'Status':>8}")
    print("-" * 150)
    
    issues = []
    
    for row in range(2, ws.max_row + 1):
        actor = ws.cell(row=row, column=3).value
        if not actor:
            continue
        
        name = ws.cell(row=row, column=2).value
        hp = ws.cell(row=row, column=4).value
        speed = ws.cell(row=row, column=5).value
        rng = ws.cell(row=row, column=6).value
        damage = ws.cell(row=row, column=7).value
        wclass = ws.cell(row=row, column=8).value
        reload = ws.cell(row=row, column=9).value
        k = ws.cell(row=row, column=11).value
        l = ws.cell(row=row, column=12).value
        m = ws.cell(row=row, column=13).value
        cost = ws.cell(row=row, column=19).value  # S = target cost
        
        if damage is None or damage == 0 or reload is None or reload == 0:
            # No weapon — compute with 0 DPS
            o, p, q, r, dps = compute_cost(hp, speed, rng, 1, 1, wclass, k, l, m)
            # Actually for no-weapon units, DPS=0
            o = (hp/100000 + speed/100 + rng*k/5 + 0) * 200 * l * m
            p = (hp*speed/25000 + 0) * l * m
            q = 0
            r = (o + p + q) / 3
            diff = r - cost
            status = "OK" if abs(diff) < 1 else "DIFF"
            print(f"{actor:<35} {hp:>8} {speed:>5} {rng:>8.3f} {damage or 0:>8} {reload or 0:>5} {k:>4} {l:>5} {m:>4} {cost:>8} {r:>10.1f} {diff:>8.1f} {status:>8}")
            if status != "OK":
                issues.append(f"  Row {row} ({actor}): R={r:.1f} S={cost} diff={diff:.1f}")
            continue
        
        o, p, q, r, dps = compute_cost(hp, speed, rng, damage, reload, wclass, k, l, m)
        diff = r - cost
        status = "OK" if abs(diff) < 1 else "DIFF"
        
        print(f"{actor:<35} {hp:>8} {speed:>5} {rng:>8.3f} {damage:>8} {reload:>5} {k:>4} {l:>5} {m:>4} {cost:>8} {r:>10.1f} {diff:>8.1f} {status:>8}")
        
        if status != "OK":
            # Also compute what Range should be
            rng_solved = solve_range(hp, speed, damage, reload, wclass, k, l, m, cost)
            issues.append(f"  Row {row} ({actor}): R={r:.1f} S={cost} diff={diff:.1f} | RangeSheet={rng:.3f} RangeSolved={rng_solved:.3f}")
    
    print(f"\n=== {len(issues)} UNBALANCED ROWS ===")
    for issue in issues:
        print(issue)

if __name__ == "__main__":
    main()
