#!/usr/bin/env python3
"""audit_balance_sheet.py — cross-reference the Cameo Armor System workbook
against in-game stats (DESIGN.md §12).

For every unit row in the workbook's unit sheets, find the in-game actor
by display name and compare:
  HP      vs Health.HP
  Speed   vs Mobile.Speed / Aircraft.Speed
  Damage  vs main weapon warhead damage x Burst        (burst rule)
  Reload  vs ReloadDelay + every inter-shot BurstDelay (burst rule)
  Range   vs weapon range in sheet units (wdist/1000, NOT cells)

Mismatches are reported for design review — the sheet is the intended
truth, but per design every mismatch is ASKED about, never auto-fixed.

Usage: python audit_balance_sheet.py [--xlsx PATH]
Default workbook: docs/design/cameo_armor_system.xlsx (repo copy)
"""

from __future__ import annotations

import argparse
import pathlib
import re
import sys

from cameo_model import Model
from report import h1, h2, table

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "tools/balance"))
import formula  # noqa: E402

DEFAULT_XLSX = str(ROOT / "docs/design/cameo_armor_system.xlsx")
UNIT_SHEETS = ("Infantry", "Tanks", "Vehicles", "Aircraft", "Defenses")
MEME_EXEMPT = {"nanoartilleryag", "nanosmokeag", "hammerheadartillery"}
UTILITY = {"genericc4", "defusekit", "leechdisinfect", "repair", "heal",
           "medikit", "dogjaw"}


def sheet_range(v: str | None) -> float | None:
    """Game range -> sheet units. The sheet stores wdist/1000 (NOT cells:
    a cell is 1024 wdist, so in-game 5c0 = 5120 wdist = sheet 5.12)."""
    value = formula.wdist_value(v)
    return None if value is None else value / 1000


def norm(name: str) -> str:
    return re.sub(r"\s+", " ", str(name)).strip().lower()


def load_fluent_names(root: pathlib.Path) -> dict[str, str]:
    names = {}
    for p in (root / "mods/cameo").rglob("*.ftl"):
        key = None
        for ln in p.read_text(encoding="utf-8-sig", errors="replace").split("\n"):
            mo = re.match(r"^([\w-]+)\s*=", ln)
            if mo:
                key = mo.group(1)
            mo = re.match(r"\s+\.name\s*=\s*(.+)", ln)
            if mo and key:
                names[key + ".name"] = mo.group(1).strip()
    return names


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    args = ap.parse_args()
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed (pip install openpyxl)")
        return 1

    m = Model()
    rs = m.rs
    root = pathlib.Path(__file__).resolve().parents[2]
    fluent = load_fluent_names(root)

    # display name -> actors (buildable roster, all factions)
    by_name: dict[str, set[str]] = {}
    seen = set()
    for fac in sorted(f.internal for f in m.real_factions()):
        for lname in m.buildable_roster(fac):
            if lname in seen:
                continue
            seen.add(lname)
            res = rs.resolve(lname)
            if res is None:
                continue
            tt = res.get("Tooltip", "Name")
            if not tt:
                continue
            if tt in fluent:
                tt = fluent[tt]
            by_name.setdefault(norm(tt), set()).add(lname)

    def game_stats(lname):
        """Sheet convention: Damage = SUM over the BASELINE loadout — all
        armaments named 'primary' whose condition is absent or negated
        (upgrade-gated weapons excluded); the GDI Battle Tank's cannon
        8000 + missiles 8000 = sheet 16000 is the reference case.
        Reload/Range come from the first (main) armament; a differing
        reload across armaments is flagged for case review."""
        res = rs.resolve(lname)
        hp = res.get("Health", "HP")
        spd = res.get("Mobile", "Speed") or res.get("Aircraft", "Speed")
        total_dmg = 0
        main = None
        reloads = set()
        for arm in res.children_named("Armament"):
            w = arm.get("Weapon")
            if not w or w.lower() in UTILITY:
                continue
            rc = arm.get("RequiresCondition") or ""
            if rc and not rc.startswith("!"):
                continue                      # upgraded/elite variants
            name = (arm.get("Name") or "primary").lower()
            if name not in ("primary", "secondary"):
                continue                      # garrisoned/targeting/etc.
            ww = rs.resolve_weapon(w)
            if ww is None:
                continue
            dmg = 0
            for c in ww.children:
                kl = c.key.lower()
                if not kl.startswith("warhead") or "friendlyfire" in kl:
                    continue
                if (c.value or "") == "HealthPercentageDamage":
                    continue
                try:
                    d = int(c.get("Damage") or 0)
                except ValueError:
                    continue
                if d > 0:
                    dmg += d
            if dmg <= 0:
                continue
            burst = int(ww.get("Burst") or 1)
            bdel = ww.get("BurstDelays") or ww.get("BurstDelay")
            reload_ = int(ww.get("ReloadDelay") or 40)
            eff_reload = formula.eff_reload(reload_, burst, bdel)
            reloads.add(eff_reload)
            total_dmg += dmg * burst
            if main is None:
                main = (eff_reload, sheet_range(ww.get("Range")))
        best = None
        if total_dmg and main:
            best = (total_dmg, main[0], main[1], len(reloads) > 1)
        return (int(hp) if hp else None,
                int(spd) if spd else None, best)

    print(h1("Balance sheet cross-reference (DESIGN.md §12)"))
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    for sheet in UNIT_SHEETS:
        ws = wb[sheet]
        rows_mis, unmatched, ambiguous, matched = [], [], [], 0
        for r in range(3, ws.max_row + 1):
            name = ws.cell(r, 2).value
            hp = ws.cell(r, 4).value
            if not name or not isinstance(hp, (int, float)):
                continue
            if str(ws.cell(r, 4).value) == "HP":
                continue                      # section header
            spd = ws.cell(r, 5).value
            rng = ws.cell(r, 6).value
            dmg = ws.cell(r, 7).value
            rel = ws.cell(r, 9).value
            key = norm(name)
            actors = by_name.get(key, set())
            if not actors:
                unmatched.append(str(name))
                continue
            if len(actors) > 1:
                ambiguous.append(f"{name} -> {', '.join(sorted(actors))}")
                continue
            actor = next(iter(actors))
            ghp, gspd, weap = game_stats(actor)
            matched += 1
            diffs = []
            if ghp is not None and abs(ghp - hp) > 1:
                diffs.append(f"HP {hp} vs game {ghp}")
            if isinstance(spd, (int, float)) and gspd is not None \
                    and abs(gspd - spd) > 1:
                diffs.append(f"Speed {spd} vs game {gspd}")
            if weap:
                gdmg, grel, grng, mixed_reload = weap
                if isinstance(dmg, (int, float)) and dmg and                         abs(gdmg - dmg) / max(dmg, 1) > 0.02:
                    diffs.append(f"Damage {dmg} vs game {gdmg}")
                if isinstance(rel, (int, float)) and abs(grel - rel) > 1:
                    diffs.append(f"Reload {rel} vs game {grel:g}"
                                 + (" [multi-reload]" if mixed_reload else ""))
                if isinstance(rng, (int, float)) and grng is not None \
                        and abs(grng - rng) > 0.15:
                    diffs.append(f"Range {rng:.2f} vs game {grng:.2f}")
            if diffs:
                rows_mis.append([str(name), actor, "; ".join(diffs)])
        print(h2(f"{sheet}: {matched} matched, {len(rows_mis)} mismatched, "
                 f"{len(unmatched)} unmatched, {len(ambiguous)} ambiguous"))
        print(table(["sheet unit", "actor", "mismatches"], rows_mis))
        if unmatched:
            print(f"unmatched: {', '.join(unmatched[:30])}"
                  + (" ..." if len(unmatched) > 30 else "") + "\n")
        if ambiguous:
            print("ambiguous: " + " | ".join(ambiguous[:10]) + "\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
