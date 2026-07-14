#!/usr/bin/env python3
"""
Comprehensive comparison of CABAL workbook vs yaml stats.
Reports mismatches for HP, Speed, Cost (from rules yaml) and Range, Damage, ReloadDelay (from weapons yaml).
"""
import openpyxl
import re
import os

MOD_ROOT = os.path.join(os.path.dirname(__file__), "..", "..", "mods", "cameo")
CABAL_DIR = os.path.join(MOD_ROOT, "ContentPacks", "TiberianSun", "CABAL")

def read_sheet():
    wb = openpyxl.load_workbook('docs/design/cameo_armor_system.xlsx', data_only=False)
    ws = wb['CABAL']
    units = {}
    for row in range(2, ws.max_row + 1):
        actor = ws.cell(row=row, column=3).value
        if not actor:
            continue
        units[actor] = {
            'name': ws.cell(row=row, column=2).value,
            'hp': ws.cell(row=row, column=4).value,
            'speed': ws.cell(row=row, column=5).value,
            'range': ws.cell(row=row, column=6).value,  # wdist/1000
            'damage': ws.cell(row=row, column=7).value,
            'reload': ws.cell(row=row, column=9).value,
            'cost': ws.cell(row=row, column=19).value,
            'k': ws.cell(row=row, column=11).value,
            'l': ws.cell(row=row, column=12).value,
            'm': ws.cell(row=row, column=13).value,
            'row': row,
        }
    return units

def read_yaml_rules():
    """Read infantry.yaml, vehicles.yaml, aircraft.yaml for HP, Speed, Cost, and weapon names."""
    actors = {}
    for fname in ['infantry.yaml', 'vehicles.yaml', 'aircraft.yaml']:
        fpath = os.path.join(CABAL_DIR, 'rules', fname)
        if not os.path.exists(fpath):
            continue
        with open(fpath, 'r', encoding='utf-8') as f:
            content = f.read()
        
        current_actor = None
        for line in content.split('\n'):
            if re.match(r'^[a-z][a-z0-9_]*:', line):
                current_actor = line.split(':')[0].strip()
                actors[current_actor] = {'file': fname, 'weapons': []}
            elif current_actor:
                m = re.match(r'^\s+HP:\s*(\d+)', line)
                if m and 'hp' not in actors[current_actor]:
                    actors[current_actor]['hp'] = int(m.group(1))
                m = re.match(r'^\s+Speed:\s*(\d+)', line)
                if m and 'speed' not in actors[current_actor]:
                    actors[current_actor]['speed'] = int(m.group(1))
                m = re.match(r'^\s+Cost:\s*(\d+)', line)
                if m and 'cost' not in actors[current_actor]:
                    actors[current_actor]['cost'] = int(m.group(1))
                m = re.match(r'^\s+Weapon:\s*(\S+)', line)
                if m:
                    actors[current_actor]['weapons'].append(m.group(1))
    return actors

def read_yaml_weapons():
    """Read CABAL weapons.yaml and shared weapon files for Range, Damage, ReloadDelay, Burst."""
    weapons = {}
    
    # CABAL weapons
    fpath = os.path.join(CABAL_DIR, 'weapons', 'weapons.yaml')
    if os.path.exists(fpath):
        with open(fpath, 'r', encoding='utf-8') as f:
            content = f.read()
        current_weapon = None
        for line in content.split('\n'):
            if re.match(r'^[A-Za-z\^][A-Za-z0-9_]*:', line) and not line.startswith(' '):
                current_weapon = line.split(':')[0].strip()
                if current_weapon not in weapons:
                    weapons[current_weapon] = {}
            elif current_weapon:
                m = re.match(r'^\s+Range:\s*(\d+)', line)
                if m and 'range' not in weapons[current_weapon]:
                    weapons[current_weapon]['range'] = int(m.group(1))
                m = re.match(r'^\s+ReloadDelay:\s*(\d+)', line)
                if m and 'reload' not in weapons[current_weapon]:
                    weapons[current_weapon]['reload'] = int(m.group(1))
                m = re.match(r'^\s+Burst:\s*(\d+)', line)
                if m:
                    weapons[current_weapon]['burst'] = int(m.group(1))
                m = re.match(r'^\s+BurstDelays:\s*(\d+)', line)
                if m:
                    weapons[current_weapon]['burst_delay'] = int(m.group(1))
                m = re.match(r'^\s+Damage:\s*(\d+)', line)
                if m and 'damage' not in weapons[current_weapon]:
                    weapons[current_weapon]['damage'] = int(m.group(1))
    
    # Shared weapons
    for wpath in [
        os.path.join(MOD_ROOT, 'weapons', 'weapons.yaml'),
        os.path.join(MOD_ROOT, 'ContentPacks', 'TiberianSun', 'Shared', 'weapons', 'weapons.yaml'),
    ]:
        if not os.path.exists(wpath):
            continue
        with open(wpath, 'r', encoding='utf-8') as f:
            content = f.read()
        current_weapon = None
        for line in content.split('\n'):
            if re.match(r'^[A-Za-z\^][A-Za-z0-9_]*:', line) and not line.startswith(' '):
                current_weapon = line.split(':')[0].strip()
                if current_weapon not in weapons:
                    weapons[current_weapon] = {}
            elif current_weapon:
                m = re.match(r'^\s+Range:\s*(\d+)', line)
                if m and 'range' not in weapons.get(current_weapon, {}):
                    weapons[current_weapon]['range'] = int(m.group(1))
                m = re.match(r'^\s+ReloadDelay:\s*(\d+)', line)
                if m and 'reload' not in weapons.get(current_weapon, {}):
                    weapons[current_weapon]['reload'] = int(m.group(1))
                m = re.match(r'^\s+Burst:\s*(\d+)', line)
                if m:
                    weapons[current_weapon]['burst'] = int(m.group(1))
                m = re.match(r'^\s+BurstDelays:\s*(\d+)', line)
                if m:
                    weapons[current_weapon]['burst_delay'] = int(m.group(1))
                m = re.match(r'^\s+Damage:\s*(\d+)', line)
                if m and 'damage' not in weapons.get(current_weapon, {}):
                    weapons[current_weapon]['damage'] = int(m.group(1))
    
    return weapons

def main():
    sheet = read_sheet()
    yaml_actors = read_yaml_rules()
    yaml_weapons = read_yaml_weapons()
    
    mismatches = []
    
    for actor, s in sorted(sheet.items()):
        y = yaml_actors.get(actor, {})
        
        # Compare HP
        s_hp = s['hp']
        y_hp = y.get('hp')
        if y_hp is not None and s_hp != y_hp:
            mismatches.append(f"  {actor}: HP sheet={s_hp} yaml={y_hp}")
        
        # Compare Speed
        s_speed = s['speed']
        y_speed = y.get('speed')
        if y_speed is not None and s_speed != y_speed:
            mismatches.append(f"  {actor}: Speed sheet={s_speed} yaml={y_speed}")
        
        # Compare Cost
        s_cost = s['cost']
        y_cost = y.get('cost')
        if y_cost is not None and s_cost != y_cost:
            mismatches.append(f"  {actor}: Cost sheet={s_cost} yaml={y_cost}")
        
        # Compare weapon stats (Range, Damage, ReloadDelay)
        actor_weapons = y.get('weapons', [])
        if actor_weapons and s['damage'] and s['damage'] > 0:
            primary_weapon = actor_weapons[0]
            w = yaml_weapons.get(primary_weapon, {})
            
            s_range_wdist = int(s['range'] * 1000) if s['range'] else 0
            y_range = w.get('range')
            if y_range is not None and s_range_wdist != y_range:
                mismatches.append(f"  {actor} ({primary_weapon}): Range sheet={s_range_wdist} yaml={y_range}")
            
            s_damage = s['damage']
            y_damage = w.get('damage')
            if y_damage is not None and s_damage != y_damage:
                mismatches.append(f"  {actor} ({primary_weapon}): Damage sheet={s_damage} yaml={y_damage}")
            
            s_reload = s['reload']
            y_reload = w.get('reload')
            if y_reload is not None and s_reload != y_reload:
                mismatches.append(f"  {actor} ({primary_weapon}): Reload sheet={s_reload} yaml={y_reload}")
    
    if mismatches:
        print(f"=== {len(mismatches)} MISMATCHES FOUND ===")
        for m in mismatches:
            print(m)
    else:
        print("No mismatches found!")

if __name__ == "__main__":
    main()
