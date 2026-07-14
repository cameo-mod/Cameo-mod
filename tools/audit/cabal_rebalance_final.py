#!/usr/bin/env python3
"""
CABAL Rebalance - Final Pass
Computes correct sheet Damage (main SpreadDamage only, no FriendlyFire/Extra/DamagesConcrete)
and ReloadDelay (with burst rule), correct K/L/M, and solves for Range.
Outputs a complete change plan for both sheet and yaml.
"""
import openpyxl
import re
import os
import json

MOD_ROOT = os.path.join(os.path.dirname(__file__), "..", "..", "mods", "cameo")
CABAL_DIR = os.path.join(MOD_ROOT, "ContentPacks", "TiberianSun", "CABAL")

# ============================================================
# Weapon template class mapping
# ============================================================
TEMPLATE_CLASS = {
    '^SmallArms': 0.75, '^Chaingun': 0.75,
    '^SniperWeapon': 0.75,
    '^LightMissile': 0.75, '^MediumMissile': 1.0, '^HeavyMissile': 1.25,
    '^MediumCannon': 1.0, '^HeavyCannon': 1.25,
    '^LaserWeapon': 1.0,
    '^TeslaWeapon': 1.25,
    '^SwordWeapon': 0.75,
    '^LightFlameWeapon': 0.75, '^MediumFlameWeapon': 1.0, '^HeavyFlameWeapon': 1.25,
    '^LightChemicalWeapon': 0.75, '^MediumChemicalWeapon': 1.0, '^HeavyChemicalWeapon': 1.25,
    '^RailgunWeapon': 1.25,
    '^MagicWeapon': 1.25,
    '^HeavyMG': 0.75, '^LightMG': 0.75,
}

# ============================================================
# L mapping from DESIGN.md
# ============================================================
L_MAP = {
    'MainBattleTankTemplate': 1.0,
    'HighTechTankTemplate': 1.0,
    'SupportVehicleTemplate': 1.25,
    'FireSupportTemplate': 1.0,
    'ArtilleryTemplate': 0.5,
    'EpicVehicleTemplate': 0.3,
    'EpicAirUnitTemplate': 0.3,
    'HelicopterTemplate': 1.0,
    'SpaceshipTemplate': 1.0,
    'BomberTemplate': 1.0,
    'FlyingInfantryTemplate': 1.0,
    'UnarmedTransportHelicopterTemplate': 1.0,
    # Infantry templates
    'HeavyInfantryTemplate': 0.8,
    'HeroInfantryTemplate': 1.0,
    'AntiTankAntiAirInfantryTemplate': 0.5,
    'LineBreakerTemplate': 0.8,
    'MechanicTemplate': 0.5,
    'ScoutInfantryTemplate': 0.5,
    'GrenadierInfantryTemplate': 0.4,
    'MortarInfantryTemplate': 0.6,
    'MeleeInfantryTemplate': 0.75,
    'SniperInfantryTemplate': 0.75,
}

# ============================================================
# Read weapons
# ============================================================
def read_weapons(fpath):
    weapons = {}
    if not os.path.exists(fpath):
        return weapons
    with open(fpath, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    current = None
    for i, line in enumerate(lines):
        if re.match(r'^[A-Za-z\^][A-Za-z0-9_]*:', line) and not line.startswith(' '):
            current = line.split(':')[0].strip()
            if current not in weapons:
                weapons[current] = {
                    'range': None, 'reload': None, 'burst': 1,
                    'burst_delays': [], 'inherits': [],
                    'warheads': [],  # list of (name, type, damage, is_friendly_fire, is_extra)
                    'target_damages': [],
                }
        elif current:
            m = re.match(r'^\s+Inherits\S*:\s*(\S+)', line)
            if m:
                weapons[current]['inherits'].append(m.group(1))
            m = re.match(r'^\s+Range:\s*(\d+)', line)
            if m and weapons[current]['range'] is None:
                weapons[current]['range'] = int(m.group(1))
            m = re.match(r'^\s+ReloadDelay:\s*(\d+)', line)
            if m and weapons[current]['reload'] is None:
                weapons[current]['reload'] = int(m.group(1))
            m = re.match(r'^\s+Burst:\s*(\d+)', line)
            if m:
                weapons[current]['burst'] = int(m.group(1))
            m = re.match(r'^\s+BurstDelays:\s*(.+)', line)
            if m:
                weapons[current]['burst_delays'] = [int(x.strip()) for x in m.group(1).split(',')]
            
            m = re.match(r'^\s+Warhead@([^:]+):\s*(\S+)', line)
            if m:
                wh_name = m.group(1)
                wh_type = m.group(2)
                dmg = None
                for j in range(i+1, min(i+5, len(lines))):
                    dm = re.match(r'^\s+Damage:\s*(\d+)', lines[j])
                    if dm:
                        dmg = int(dm.group(1))
                        break
                is_ff = 'FriendlyFire' in wh_name
                is_extra = 'Extra' in wh_name
                weapons[current]['warheads'].append({
                    'name': wh_name, 'type': wh_type, 'damage': dmg,
                    'is_friendly_fire': is_ff, 'is_extra': is_extra,
                })
                # Also track TargetDamage as a damage source
                if wh_type == 'TargetDamage' and dmg:
                    weapons[current]['target_damages'].append(dmg)
    return weapons

def compute_weapon_stats(wname, all_weapons, depth=0):
    """Compute sheet Damage, ReloadDelay, Range, WeaponClass from weapon definition."""
    if wname not in all_weapons or depth > 10:
        return None
    w = all_weapons[wname]
    
    # Sheet Damage = sum of main SpreadDamage (not FriendlyFire, not Extra) x burst
    # Also include TargetDamage if no SpreadDamage (e.g. hacker weapons)
    main_spread = 0
    target_dmg_total = 0
    template_classes = []
    
    for wh in w['warheads']:
        if wh['type'] == 'SpreadDamage' and not wh['is_friendly_fire'] and not wh['is_extra'] and wh['damage']:
            main_spread += wh['damage']
        if wh['type'] == 'TargetDamage' and wh['damage']:
            target_dmg_total += wh['damage']
    
    # Collect template classes from inherits
    for inh in w.get('inherits', []):
        if inh in TEMPLATE_CLASS:
            template_classes.append(TEMPLATE_CLASS[inh])
    
    # Also resolve non-template parents for warheads they contribute
    for inh in w.get('inherits', []):
        if inh not in TEMPLATE_CLASS and inh in all_weapons:
            parent = compute_weapon_stats(inh, all_weapons, depth+1)
            if parent:
                # If this weapon doesn't override warheads, parent's damage counts
                if main_spread == 0 and parent['main_spread'] > 0:
                    main_spread = parent['main_spread']
                if target_dmg_total == 0 and parent.get('target_dmg', 0) > 0:
                    target_dmg_total = parent['target_dmg']
                # Inherit template classes from parent's inherits too
                for pinh in all_weapons.get(inh, {}).get('inherits', []):
                    if pinh in TEMPLATE_CLASS and pinh not in [i for i in w.get('inherits', []) if i in TEMPLATE_CLASS]:
                        template_classes.append(TEMPLATE_CLASS[pinh])
    
    # WeaponClass = average of template classes
    if template_classes:
        wclass = sum(template_classes) / len(template_classes)
    else:
        wclass = 1.0
    
    burst = w['burst']
    burst_delays = w['burst_delays']
    reload = w['reload'] or 0
    
    # Sheet Damage = (main_spread + target_dmg) x burst
    # If no SpreadDamage, use TargetDamage
    if main_spread > 0:
        sheet_damage = main_spread * burst
    elif target_dmg_total > 0:
        sheet_damage = target_dmg_total * burst
    else:
        sheet_damage = 0
    
    # Sheet ReloadDelay = reload + (burst-1) x burst_delay
    if burst > 1 and burst_delays:
        sheet_reload = reload + (burst - 1) * burst_delays[0]
    else:
        sheet_reload = reload
    
    return {
        'range': w['range'],
        'reload': sheet_reload,
        'damage': sheet_damage,
        'wclass': wclass,
        'burst': burst,
        'main_spread': main_spread,
        'target_dmg': target_dmg_total,
    }

# ============================================================
# Read rules
# ============================================================
def read_rules():
    actors = {}
    for fname in ['infantry.yaml', 'vehicles.yaml', 'aircraft.yaml']:
        fpath = os.path.join(CABAL_DIR, 'rules', fname)
        if not os.path.exists(fpath):
            continue
        with open(fpath, 'r', encoding='utf-8') as f:
            content = f.read()
        
        current = None
        for line in content.split('\n'):
            if re.match(r'^[a-z][a-z0-9_]*:', line):
                current = line.split(':')[0].strip()
                actors[current] = {
                    'file': fname, 'weapons': [], 'prerequisites': [],
                    'inherits': [], 'hp': None, 'speed': None, 'cost': None,
                }
            elif current:
                m = re.match(r'^\s+Weapon:\s*(\S+)', line)
                if m:
                    actors[current]['weapons'].append(m.group(1))
                m = re.match(r'^\s+HP:\s*(\d+)', line)
                if m and actors[current]['hp'] is None:
                    actors[current]['hp'] = int(m.group(1))
                m = re.match(r'^\s+Speed:\s*(\d+)', line)
                if m and actors[current]['speed'] is None:
                    actors[current]['speed'] = int(m.group(1))
                m = re.match(r'^\s+Cost:\s*(\d+)', line)
                if m and actors[current]['cost'] is None:
                    actors[current]['cost'] = int(m.group(1))
                m = re.match(r'^\s+Prerequisites:\s*(.+)', line)
                if m:
                    actors[current]['prerequisites'] = [x.strip() for x in m.group(1).split(',')]
                m = re.match(r'^\s+Inherits\S*:\s*(\S+)', line)
                if m:
                    actors[current]['inherits'].append(m.group(1))
    return actors

# ============================================================
# K/L/M determination
# ============================================================
def determine_L(actor_data):
    # Epic templates always override
    for inh in actor_data.get('inherits', []):
        if 'Epic' in inh:
            for key, val in L_MAP.items():
                if key in inh:
                    return val
    # Otherwise pick the highest L among matching templates
    best_l = None
    for inh in actor_data.get('inherits', []):
        for key, val in L_MAP.items():
            if key in inh:
                if best_l is None or val > best_l:
                    best_l = val
    return best_l if best_l is not None else 1.0

def determine_M(actor_data):
    prereqs = actor_data.get('prerequisites', [])
    # Epic: M=1.0
    for inh in actor_data.get('inherits', []):
        if 'Epic' in inh:
            return 1.0
    # Check tech tier from prerequisites
    has_core = any('cabal_core' in p for p in prereqs)
    has_techcenter = any('techcenter' in p for p in prereqs)
    has_radar = any('radar' in p for p in prereqs)
    has_helipad = any('helipad' in p for p in prereqs)
    has_warfactory = any('mechfactory' in p or 'warfactory' in p for p in prereqs)
    has_cyborgfactory = any('cyborgfactory' in p for p in prereqs)
    
    if has_core:
        return 0.5  # T4/5 (core requires techcenter)
    if has_techcenter:
        return 0.75  # T3
    # T1/T2 = 1.0
    return 1.0

def determine_K(actor_data):
    k = 1.0
    inherits = actor_data.get('inherits', [])
    weapons = actor_data.get('weapons', [])
    all_str = str(inherits) + str(weapons)
    
    # EMP
    if 'FrontalEMP' in all_str or 'EMP' in all_str:
        k += 0.25
    # Cloak/stealth
    if 'Cloak' in all_str or 'Stealth' in all_str or 'stealth' in all_str:
        k += 0.25
    # Mind control / hacking
    if 'HackingControl' in all_str or 'MindControl' in all_str:
        k += 0.25
    # Trap
    if any('Trap' in w for w in weapons):
        k += 0.25
    # Vampire heal-on-attack (Dissolver)
    if 'Dissolver' in str(actor_data):
        k += 0.25
    
    return min(k, 2.0)

# ============================================================
# Balance formula
# ============================================================
def solve_range(hp, speed, damage, reload, wclass, k, l, m, cost):
    if damage == 0 or reload == 0:
        return 0
    dps = damage / reload * wclass
    h = hp / 100000
    s = speed / 100
    d = dps / 200
    p_const = hp * speed / 25000
    q_coeff = hp * speed * dps / 12500000
    
    lhs = 3 * cost / (l * m)
    constant = 200 * (h + s + d) + p_const
    range_coeff = k * (40 + dps / 2.5 + q_coeff)
    
    if range_coeff == 0:
        return 0
    return (lhs - constant) / range_coeff

def compute_cost(hp, speed, rng, damage, reload, wclass, k, l, m):
    if damage == 0 or reload == 0:
        dps = 0
    else:
        dps = damage / reload * wclass
    o = (hp/100000 + speed/100 + rng*k/5 + dps/200) * 200 * l * m
    p = (hp*speed/25000 + rng*k*dps/2.5) * l * m
    q = hp*speed*rng*k*dps*l*m / 12500000
    return (o + p + q) / 3, dps

# ============================================================
# Main
# ============================================================
def main():
    # Read all weapons
    all_weapons = {}
    for fpath in [
        os.path.join(CABAL_DIR, 'weapons', 'weapons.yaml'),
        os.path.join(MOD_ROOT, 'weapons', 'weapons.yaml'),
        os.path.join(MOD_ROOT, 'ContentPacks', 'TiberianSun', 'Shared', 'weapons', 'weapons.yaml'),
    ]:
        all_weapons.update(read_weapons(fpath))
    
    # Read rules
    actors = read_rules()
    
    # Read sheet
    wb = openpyxl.load_workbook('docs/design/cameo_armor_system.xlsx', data_only=False)
    ws = wb['CABAL']
    
    sheet_rows = {}
    for row in range(2, ws.max_row + 1):
        actor = ws.cell(row=row, column=3).value
        if actor:
            sheet_rows[actor] = row
    
    print("=== CABAL REBALANCE - CORRECT VALUES ===")
    print()
    
    results = []
    
    for actor, row in sorted(sheet_rows.items(), key=lambda x: x[1]):
        s_hp = ws.cell(row=row, column=4).value
        s_speed = ws.cell(row=row, column=5).value
        s_rng = ws.cell(row=row, column=6).value
        s_dmg = ws.cell(row=row, column=7).value
        s_wclass = ws.cell(row=row, column=8).value
        s_reload = ws.cell(row=row, column=9).value
        s_k = ws.cell(row=row, column=11).value
        s_l = ws.cell(row=row, column=12).value
        s_m = ws.cell(row=row, column=13).value
        s_cost = ws.cell(row=row, column=19).value
        
        a = actors.get(actor, {})
        
        # Correct K/L/M
        k = determine_K(a)
        l = determine_L(a)
        m = determine_M(a)
        
        # Weapon stats from yaml - pick weapon with highest DPS
        actor_weapons = a.get('weapons', [])
        primary_weapon = None
        wstats = None
        best_dps = 0
        for wname in actor_weapons:
            ws_stats = compute_weapon_stats(wname, all_weapons)
            if ws_stats and ws_stats['damage'] > 0 and ws_stats['reload'] > 0:
                w_dps = ws_stats['damage'] / ws_stats['reload'] * ws_stats['wclass']
                if w_dps > best_dps:
                    best_dps = w_dps
                    primary_weapon = wname
                    wstats = ws_stats
        
        if wstats:
            damage = wstats['damage']
            reload = wstats['reload']
            wclass = wstats['wclass']
            yaml_range = wstats['range']
        else:
            damage = 0
            reload = 0
            wclass = s_wclass or 1.0
            yaml_range = 0
        
        hp = a.get('hp') or s_hp
        speed = a.get('speed') or s_speed
        cost = s_cost
        
        # Solve for Range
        rng_solved = solve_range(hp, speed, damage, reload, wclass, k, l, m, cost)
        r_calc, dps = compute_cost(hp, speed, rng_solved, damage, reload, wclass, k, l, m)
        
        # Classify range
        range_status = "OK"
        if rng_solved < 0:
            range_status = "NEGATIVE"
        elif rng_solved > 15:
            range_status = "ABSURD"
        elif rng_solved > 10:
            range_status = "HIGH"
        elif rng_solved < 0.5 and damage > 0:
            range_status = "TOO_LOW"
        
        changes = []
        if s_k != k: changes.append(f"K {s_k}->{k}")
        if s_l != l: changes.append(f"L {s_l}->{l}")
        if s_m != m: changes.append(f"M {s_m}->{m}")
        if s_dmg != damage: changes.append(f"Dmg {s_dmg}->{damage}")
        if s_reload != reload: changes.append(f"Rld {s_reload}->{reload}")
        if s_wclass != wclass: changes.append(f"WC {s_wclass}->{wclass:.3f}")
        if s_hp != hp: changes.append(f"HP {s_hp}->{hp}")
        if s_speed != speed: changes.append(f"Spd {s_speed}->{speed}")
        
        # Check formula
        o_formula = ws.cell(row=row, column=15).value
        formula_broken = False
        if o_formula and isinstance(o_formula, str) and o_formula.startswith('='):
            if f"D{row}" not in o_formula:
                changes.append("FORMULA BROKEN")
                formula_broken = True
        
        yaml_cost = a.get('cost')
        if yaml_cost and s_cost != yaml_cost:
            changes.append(f"Cost sheet={s_cost} yaml={yaml_cost}")
        
        print(f"Row {row:2d} {actor:<35} [{range_status}]")
        if changes:
            for c in changes:
                print(f"       {c}")
        print(f"       HP={hp} Spd={speed} Dmg={damage} Rld={reload} WC={wclass:.3f} K={k} L={l} M={m} Cost={cost}")
        print(f"       RngSolved={rng_solved:.3f} (wdist={int(rng_solved*1000)}) R={r_calc:.1f} DPS={dps:.1f}")
        if primary_weapon:
            print(f"       Weapon={primary_weapon} YamlRange={yaml_range}")
        print()
        
        results.append({
            'actor': actor, 'row': row,
            'hp': hp, 'speed': speed, 'damage': damage, 'reload': reload,
            'wclass': wclass, 'k': k, 'l': l, 'm': m, 'cost': cost,
            'range': rng_solved, 'yaml_range': yaml_range,
            'range_status': range_status, 'changes': changes,
            'formula_broken': formula_broken,
            'primary_weapon': primary_weapon,
            'yaml_cost': yaml_cost,
        })
    
    # Summary
    absurd = [r for r in results if r['range_status'] in ('ABSURD', 'NEGATIVE', 'TOO_LOW')]
    high = [r for r in results if r['range_status'] == 'HIGH']
    ok = [r for r in results if r['range_status'] == 'OK']
    no_weapon = [r for r in results if r['range_status'] == 'OK' and r['damage'] == 0]
    
    print(f"=== SUMMARY ===")
    print(f"  OK range: {len(ok) - len(no_weapon)} units (+ {len(no_weapon)} no weapon)")
    print(f"  HIGH range (>10): {len(high)} units: {[r['actor'] for r in high]}")
    print(f"  ABSURD/NEGATIVE/TOO_LOW: {len(absurd)} units: {[r['actor'] for r in absurd]}")
    print(f"  Formula broken: {sum(1 for r in results if r['formula_broken'])} rows")

if __name__ == "__main__":
    main()
