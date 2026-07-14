#!/usr/bin/env python3
"""Read all CABAL rows from the workbook and print their stats."""
import openpyxl

wb = openpyxl.load_workbook('docs/design/cameo_armor_system.xlsx', data_only=False)

# Print column headers for each sheet, then find CABAL rows
for sheet_name in ['Infantry', 'Tanks', 'Vehicles', 'Aircraft', 'Defenses']:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    
    # Print header row (row 1)
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value is not None:
            headers[col] = str(cell.value)
    print(f"  Columns: {headers}")
    
    # Find CABAL rows
    for row in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=2).value  # column B = name
        if name_cell and 'cabal' in str(name_cell).lower():
            row_data = {}
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    col_name = headers.get(col, f"col{col}")
                    row_data[col_name] = cell.value
            print(f"  Row {row}: {row_data}")

# Also check if there's a separate CABAL sheet
if 'CABAL' in wb.sheetnames:
    ws = wb['CABAL']
    print(f"\n=== CABAL (dedicated sheet) ===")
    for row in range(1, ws.max_row + 1):
        row_data = []
        has_content = False
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                has_content = True
                row_data.append(f"{cell.coordinate}={cell.value}")
        if has_content:
            print(f"  Row {row}: {' | '.join(row_data)}")
