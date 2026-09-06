"""The generated balance workbook must import as a no-op without column drift."""

from __future__ import annotations

import copy
import pathlib
import sys
import unittest

import _bootstrap  # noqa: F401 — sys.path side effect

ROOT = pathlib.Path(__file__).resolve().parents[2]
BALANCE = ROOT / "tools" / "balance"
if str(BALANCE) not in sys.path:
    sys.path.insert(0, str(BALANCE))

try:
    import openpyxl
except ModuleNotFoundError:  # The normal game build does not install spreadsheet extras.
    openpyxl = None


@unittest.skipUnless(openpyxl is not None, "openpyxl spreadsheet test dependency")
class WorkbookRoundTripTest(unittest.TestCase):
    @staticmethod
    def fixture_unit():
        return {
            "name": "Fixture Tank",
            "hp": {"v": "125000"},
            "speed": {"v": "75"},
            "armor": {"v": "Heavy"},
            "cost": {"v": "900"},
            "firepower_multiplier": {"v": "1.25"},
            "design": {
                "subtype": "MainBattleTank", "tech_tier": None,
                "unit_class": 1.5, "special": 0.9,
            },
            "armaments": [{
                "slot": "Armament@PRIMARY", "weapon": "FixtureCannon",
                "pricing": True, "reloaddelay": "60", "burst": "3",
                "burstdelays": "2, 7", "range": "6000",
                "design_weapon_class": 1.25,
                "warheads": ["^Warhead_CannonAP_Heavy"],
                "damage_warheads": [{
                    "tag": "CannonAP_Heavy", "type": "AreaDamage",
                    "damage": "12000", "spread": "100", "falloff": "100, 0",
                }],
            }],
        }

    @staticmethod
    def generated_sheet(build, unit, section="vehicles"):
        wb = openpyxl.Workbook()
        ws = wb.active
        for column, header in enumerate(build.HDR, 1):
            ws.cell(row=1, column=column, value=header)
        build.build._anchors = {}
        build.build._tier_map = {"fixture_tank": 0.8}
        build.unit_rows(ws, "fixture", "fixture_tank", unit, section, 2)
        return ws

    def test_fresh_generator_row_imports_without_mutating_ledger_data(self):
        import build_workbook as build
        import import_workbook as importer

        unit = self.fixture_unit()
        ws = self.generated_sheet(build, unit)

        before = copy.deepcopy(unit)
        touched, changes, changed_actors = importer.import_sheet(
            ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})

        self.assertFalse(touched)
        self.assertEqual(changes, 0)
        self.assertEqual(changed_actors, set())
        self.assertEqual(unit, before)

    def test_editing_displayed_weapon_defaults_creates_missing_raw_fields(self):
        import build_workbook as build
        import import_workbook as importer

        unit = self.fixture_unit()
        del unit["armaments"][0]["burst"]
        del unit["armaments"][0]["reloaddelay"]
        del unit["armaments"][0]["range"]
        ws = self.generated_sheet(build, unit)
        weapon_row = 3
        self.assertEqual(
            ws.cell(weapon_row, build.COL["Reload"]).value,
            build.formula.ENGINE_DEFAULT_RELOAD_DELAY)
        self.assertEqual(
            ws.cell(weapon_row, build.COL["Burst"]).value,
            build.formula.ENGINE_DEFAULT_BURST)
        self.assertEqual(
            ws.cell(weapon_row, build.COL["Range(wd)"]).value,
            build.formula.ENGINE_DEFAULT_RANGE)

        before = copy.deepcopy(unit)
        touched, changes, changed_actors = importer.import_sheet(
            ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})
        self.assertFalse(touched)
        self.assertEqual(changes, 0)
        self.assertEqual(changed_actors, set())
        self.assertEqual(unit, before)

        ws.cell(weapon_row, build.COL["Reload"], value=4)
        # Keep the existing two-entry BurstDelays list valid: Burst - 1 = 2.
        ws.cell(weapon_row, build.COL["Burst"], value=3)
        ws.cell(weapon_row, build.COL["Range(wd)"], value=1000)

        touched, changes, changed_actors = importer.import_sheet(
            ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})

        self.assertTrue(touched)
        self.assertEqual(changes, 3)
        self.assertEqual(changed_actors, {"fixture_tank"})
        self.assertEqual(unit["armaments"][0]["reloaddelay"], "4")
        self.assertEqual(unit["armaments"][0]["burst"], "3")
        self.assertEqual(unit["armaments"][0]["range"], "1000")

    def test_cell_notation_range_displays_numerically_and_imports_as_no_op(self):
        import build_workbook as build
        import import_workbook as importer

        unit = self.fixture_unit()
        unit["armaments"][0]["range"] = "40c0"
        ws = self.generated_sheet(build, unit)
        weapon_row = 3
        self.assertEqual(ws.cell(weapon_row, build.COL["Range(wd)"]).value, 40960)

        before = copy.deepcopy(unit)
        touched, changes, changed_actors = importer.import_sheet(
            ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})
        self.assertFalse(touched)
        self.assertEqual(changes, 0)
        self.assertEqual(changed_actors, set())
        self.assertEqual(unit, before)

        ws.cell(weapon_row, build.COL["Range(wd)"], value="41c0")
        touched, changes, changed_actors = importer.import_sheet(
            ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})
        self.assertTrue(touched)
        self.assertEqual(changes, 1)
        self.assertEqual(changed_actors, {"fixture_tank"})
        self.assertEqual(unit["armaments"][0]["range"], "41984")

    def test_identity_columns_are_resolved_after_reorder(self):
        import build_workbook as build
        import import_workbook as importer

        unit = self.fixture_unit()
        source = self.generated_sheet(build, unit)
        headers = ["Inserted", "Name"] + [
            header for header in build.HDR if header not in {"Actor", "Name"}
        ] + ["Actor"]
        wb = openpyxl.Workbook()
        ws = wb.active
        for column, header in enumerate(headers, 1):
            ws.cell(1, column, header)
            if header == "Inserted":
                continue
            source_column = build.COL[header]
            for row in range(2, source.max_row + 1):
                ws.cell(row, column, source.cell(row, source_column).value)

        header_columns = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        ws.cell(3, header_columns["Reload"], 61)
        touched, changes, changed_actors = importer.import_sheet(
            ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})
        self.assertTrue(touched)
        self.assertEqual(changes, 1)
        self.assertEqual(changed_actors, {"fixture_tank"})
        self.assertEqual(unit["armaments"][0]["reloaddelay"], "61")

    def test_fractional_engine_cells_are_rejected_before_mutation(self):
        import build_workbook as build
        import import_workbook as importer

        for header, row, value in (
                ("HP", 2, 125000.5), ("Speed", 2, 75.5),
                ("Speed", 2, "40c0"), ("Speed", 2, 2 ** 31),
                ("Damage", 3, 12000.5),
                ("Reload", 3, 2.5), ("Burst", 3, 1.5)):
            unit = self.fixture_unit()
            before = copy.deepcopy(unit)
            ws = self.generated_sheet(build, unit)
            ws.cell(row, build.COL[header], value)
            with self.subTest(header=header):
                with self.assertRaises(ValueError):
                    importer.import_sheet(
                        ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})
                self.assertEqual(unit, before)

    def test_scalar_weapon_range_rejects_distribution_text(self):
        import build_workbook as build
        import import_workbook as importer

        unit = self.fixture_unit()
        before = copy.deepcopy(unit)
        ws = self.generated_sheet(build, unit)
        ws.cell(3, build.COL["Range(wd)"], "40c0, 42c0")
        with self.assertRaisesRegex(ValueError, "scalar WDist"):
            importer.import_sheet(
                ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})
        self.assertEqual(unit, before)

    def test_invalid_cadence_is_rejected_before_mutation(self):
        import build_workbook as build
        import import_workbook as importer

        for header, value, message in (
                ("Reload", 0, "greater than zero"),
                ("Burst", 4, "Burst - 1")):
            unit = self.fixture_unit()
            before = copy.deepcopy(unit)
            ws = self.generated_sheet(build, unit)
            ws.cell(3, build.COL[header], value)
            with self.subTest(header=header):
                with self.assertRaisesRegex(ValueError, message):
                    importer.import_sheet(
                        ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})
                self.assertEqual(unit, before)

    def test_defaulted_design_inputs_only_create_fields_after_an_edit(self):
        import build_workbook as build
        import import_workbook as importer

        unit = self.fixture_unit()
        unit["design"] = {"tech_tier": None}
        del unit["armaments"][0]["design_weapon_class"]
        ws = self.generated_sheet(build, unit)

        before = copy.deepcopy(unit)
        touched, changes, changed_actors = importer.import_sheet(
            ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})
        self.assertFalse(touched)
        self.assertEqual(changes, 0)
        self.assertEqual(changed_actors, set())
        self.assertEqual(unit, before)

        ws.cell(2, build.COL["Class"], value="MainBattleTank")
        ws.cell(2, build.COL["TechTier"], value=1)
        ws.cell(2, build.COL["UnitClass"], value=1.25)
        ws.cell(2, build.COL["Special"], value=0.9)
        ws.cell(3, build.COL["WeapClass"], value=1.5)
        touched, changes, changed_actors = importer.import_sheet(
            ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})

        self.assertTrue(touched)
        self.assertEqual(changes, 0, "design-only edits are not raw stat changes")
        self.assertEqual(changed_actors, {"fixture_tank"})
        self.assertEqual(unit["design"]["subtype"], "MainBattleTank")
        self.assertEqual(unit["design"]["tech_tier"], 1)
        self.assertEqual(unit["design"]["unit_class"], 1.25)
        self.assertEqual(unit["design"]["special"], 0.9)
        self.assertEqual(unit["armaments"][0]["design_weapon_class"], 1.5)

    def test_clearing_burst_delay_writes_the_engine_default(self):
        import build_workbook as build
        import import_workbook as importer

        unit = self.fixture_unit()
        ws = self.generated_sheet(build, unit)
        ws.cell(3, build.COL["BurstDel"]).value = None

        touched, changes, changed_actors = importer.import_sheet(
            ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})

        self.assertTrue(touched)
        self.assertEqual(changes, 1)
        self.assertEqual(changed_actors, {"fixture_tank"})
        self.assertEqual(unit["armaments"][0]["burstdelays"], "5")

    def test_clearing_an_explicit_default_burst_delay_is_a_no_op(self):
        import build_workbook as build
        import import_workbook as importer

        unit = self.fixture_unit()
        unit["armaments"][0]["burstdelays"] = "5"
        ws = self.generated_sheet(build, unit)
        ws.cell(3, build.COL["BurstDel"]).value = None

        before = copy.deepcopy(unit)
        touched, changes, changed_actors = importer.import_sheet(
            ws, {"fixture_tank": unit}, {"fixture_tank": 0.8})
        self.assertFalse(touched)
        self.assertEqual(changes, 0)
        self.assertEqual(changed_actors, set())
        self.assertEqual(unit, before)

    def test_cells_without_safe_backing_fields_stay_locked(self):
        import build_workbook as build

        unit = self.fixture_unit()
        del unit["hp"]
        del unit["speed"]
        del unit["cost"]
        unit["armaments"][0]["damage_warheads"] = []

        wb = openpyxl.Workbook()
        ws = wb.active
        for column, header in enumerate(build.HDR, 1):
            ws.cell(row=1, column=column, value=header)
        build.build._anchors = {}
        build.build._tier_map = {"fixture_tank": 0.8}
        unit_unlock, weapon_unlock = [], []
        last_row = build.write_unit(
            ws, "fixture", "fixture_tank", unit, "defenses", 2,
            unit_unlock, weapon_unlock)
        build.finish_sheet(ws, last_row, unit_unlock, weapon_unlock)

        for name in ("HP", "Speed", "Cost"):
            self.assertTrue(ws.cell(2, build.COL[name]).protection.locked)
        self.assertTrue(ws.cell(3, build.COL["Damage"]).protection.locked)
        self.assertFalse(ws.cell(3, build.COL["Burst"]).protection.locked)


if __name__ == "__main__":
    unittest.main()
