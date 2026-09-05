"""Read-only consumer migration and fail-closed prospective write boundaries."""
import copy
import json
import pathlib
import sys
import tempfile
import unittest
from unittest.mock import patch

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1] / 'balance'))
import check_band
import firepower
import fit_class
import propose_rebalance
import propose_class_rebalance as proposal
import update_ranges

try:
    import openpyxl
except ImportError:
    openpyxl = None


def unit():
    return {'hp': {'v': '10000'}, 'speed': {'v': '60'}, 'cost': {'v': '500'},
            'design': {'class_anchor': 'mbt'},
            'firepower_multiplier': {'v': .99},
            'resolved_firepower_modifiers': [{'modifier': 50, 'types': []},
                                             {'modifier': 80, 'types': ['primary']}],
            'armaments': [{'slot': 'Armament@PRIMARY', 'weapon': 'Test', 'pricing': True,
                           'range': '5000', 'reloaddelay': '10', 'burst': '1',
                           'warheads': ['^Warhead_Test'],
                           'damage_warheads': [{'type': 'AreaDamage', 'tag': 'Test', 'damage': '1000'}]}]}


class Consumers(unittest.TestCase):
    def test_upstream_class_membership_delegation_remains_shared(self):
        import class_membership
        for module in (proposal, update_ranges):
            for subtype in ('LineBreaker', 'MortarInfantry', 'Artillery', None):
                self.assertEqual(module.subtype_to_anchor(subtype), class_membership.subtype_to_anchor(subtype))

    def test_one_arm_consumers_agree(self):
        u = unit()
        self.assertEqual(fit_class.unit_inputs(u)[0][3], 40)
        self.assertEqual(check_band.unit_inputs(u)[3], 40)
        self.assertEqual(update_ranges.unit_dps(u), 40)
        self.assertEqual(propose_rebalance.unit_row(u)[3], 40)

    def test_fractional_legacy_is_not_divided_by_100(self):
        u = unit()
        del u['resolved_firepower_modifiers']
        self.assertEqual(propose_rebalance.unit_row(u)[3], 99)
        self.assertEqual(update_ranges.unit_dps(u), 99)
        self.assertEqual(check_band.unit_inputs(u)[3], 99)

    def test_zero_is_zero(self):
        u = unit()
        u['resolved_firepower_modifiers'][0]['modifier'] = 0
        self.assertIsNone(check_band.unit_inputs(u))
        self.assertEqual(update_ranges.unit_dps(u), 0)
        self.assertEqual(propose_rebalance.unit_row(u)[3], 0)

    def test_conditional_variants_and_garrison_are_not_added(self):
        u = unit()
        for extra in ({'requires': 'elite'}, {'pricing': False}, {'unresolved': True}):
            arm = copy.deepcopy(u['armaments'][0])
            arm.update(extra)
            u['armaments'].append(arm)
        self.assertEqual(check_band.unit_inputs(u)[3], 40)
        self.assertEqual(update_ranges.unit_dps(u), 40)
        self.assertEqual(propose_rebalance.unit_row(u)[3], 40)

    def test_secondary_scope_and_primary_only_report_are_explicit(self):
        u = unit()
        second = copy.deepcopy(u['armaments'][0])
        second.update(slot='Armament@SECONDARY', armament_name='secondary')
        u['armaments'].append(second)
        self.assertEqual(update_ranges.unit_dps(u), 90)
        self.assertEqual(check_band.unit_inputs(u)[3], 90)
        self.assertEqual(propose_rebalance.unit_row(u)[3], 40)

    def test_negative_default_condition_keeps_base_weapon(self):
        arm = unit()['armaments'][0]
        arm['requires'] = '!elite'
        self.assertTrue(firepower.priced_by_default(arm))

    def test_nested_band_fixture_still_supported(self):
        u = unit()
        arm = u['armaments'][0]
        stats = {k: arm.pop(k) for k in ('damage_warheads', 'range', 'reloaddelay', 'burst')}
        arm['stats'] = stats
        self.assertEqual(check_band.unit_inputs(u)[3], 40)


class WriteBoundaries(unittest.TestCase):
    def test_range_cli_rejects_before_any_ledger_write(self):
        with tempfile.TemporaryDirectory() as directory:
            root = pathlib.Path(directory)
            legacy = unit()
            del legacy['resolved_firepower_modifiers']
            for name, u in [('a', legacy), ('z', unit())]:
                (root / (name + '.json')).write_text(json.dumps(
                    {'ledger': name, 'sections': {'vehicles': {name: u}}}), encoding='utf-8')
            before = {p: p.read_bytes() for p in root.iterdir()}
            with patch.object(update_ranges, 'LEDGER_DIR', root), \
                 patch.object(update_ranges, 'load_anchors', return_value={'mbt': {'spec': {}}}), \
                 patch.object(sys, 'argv', ['update_ranges.py', '--confirm']), \
                 patch.object(update_ranges, 'process_ledger') as process:
                with self.assertRaisesRegex(SystemExit, 'no ledgers written'):
                    update_ranges.main()
                process.assert_not_called()
            self.assertEqual(before, {p: p.read_bytes() for p in root.iterdir()})

    def test_range_guard_respects_faction_filter(self):
        doc = {'ledger': 'zerg', 'sections': {'infantry': {'hydra': unit()}}}
        update_ranges.ensure_write_supported(doc, {'mbt': {}}, 'terran')
        with self.assertRaises(SystemExit):
            update_ranges.ensure_write_supported(doc, {'mbt': {}}, 'zerg')

    def test_replacement_proposal_stops_before_output(self):
        self.check_proposal_blocked(unit())

    def test_empty_resolved_list_does_not_reenable_legacy_proposal(self):
        u = unit()
        u['resolved_firepower_modifiers'] = []
        self.check_proposal_blocked(u)

    def check_proposal_blocked(self, u):
        with tempfile.TemporaryDirectory() as directory:
            root = pathlib.Path(directory)
            (root / 'test.json').write_text(json.dumps(
                {'sections': {'vehicles': {'test': u}}}), encoding='utf-8')
            spec = {'cost0': 500, 'speed0': 60}
            with patch.object(proposal, 'LEDGER_DIR', root), \
                 patch.object(proposal, 'load_anchors', return_value={'mbt': {'spec': spec}}), \
                 patch.object(proposal, 'band_for', return_value=(0, 10000)), \
                 patch.object(sys, 'argv', ['propose_class_rebalance.py', '--class', 'mbt']), \
                 patch.object(pathlib.Path, 'write_text') as write:
                with self.assertRaisesRegex(SystemExit, 'No proposal written'):
                    proposal.main()
                write.assert_not_called()


@unittest.skipUnless(openpyxl, 'openpyxl spreadsheet test dependency')
class WorkbookConsumers(unittest.TestCase):
    def test_membership_source_participates_in_workbook_fingerprint(self):
        import build_workbook as build
        import class_membership
        self.assertEqual(build.subtype_to_anchor('LineBreaker'), 'line_breaker')
        before = build.workbook_fingerprint()
        read_text = pathlib.Path.read_text
        def changed(path, *args, **kwargs):
            text = read_text(path, *args, **kwargs)
            return text + '\n# changed mapping source\n' if path == pathlib.Path(class_membership.__file__) else text
        with patch.object(pathlib.Path, 'read_text', changed):
            self.assertNotEqual(build.workbook_fingerprint(), before)

    def test_legacy_fraction_and_empty_resolved_list(self):
        import build_workbook as build
        for resolved, expected in ((False, .99), (True, 1)):
            u = unit()
            if resolved:
                u['resolved_firepower_modifiers'] = []
            else:
                del u['resolved_firepower_modifiers']
            ws = openpyxl.Workbook().active
            build.unit_rows(ws, 'test', 'test', u, 'vehicles', 2)
            self.assertEqual(ws.cell(3, build.COL['FirepowerMultiplier']).value, expected)

    def test_weapon_row_uses_own_locked_multiplier_and_import_is_noop(self):
        import build_workbook as build
        import import_workbook as importer
        u = unit()
        second = copy.deepcopy(u['armaments'][0])
        second.update(slot='Armament@SECONDARY', armament_name='secondary')
        u['armaments'].append(second)
        before = copy.deepcopy(u)
        ws = openpyxl.Workbook().active
        for i, header in enumerate(build.HDR, 1):
            ws.cell(1, i, header)
        build.build._anchors = {}
        build.build._tier_map = {}
        build.unit_rows(ws, 'test', 'test', u, 'vehicles', 2)
        for row, factor in ((3, .4), (4, .5)):
            self.assertEqual(ws.cell(row, build.COL['FirepowerMultiplier']).value, factor)
            self.assertTrue(ws.cell(row, build.COL['FirepowerMultiplier']).protection.locked)
            self.assertIn('*' + build.L('FirepowerMultiplier') + str(row),
                          ws.cell(row, build.COL['DPS']).value)
        # Even a programmatic edit to this locked diagnostic field is not imported.
        ws.cell(3, build.COL['FirepowerMultiplier'], 999)
        self.assertFalse(importer.import_sheet(ws, {'test': u})[0])
        self.assertEqual(u, before)

    def test_conditional_weapon_is_displayed_but_not_summed(self):
        import build_workbook as build
        u = unit()
        extra = copy.deepcopy(u['armaments'][0])
        extra.update(requires='elite', slot='Armament@ELITE')
        u['armaments'].append(extra)
        ws = openpyxl.Workbook().active
        build.unit_rows(ws, 'test', 'test', u, 'vehicles', 2)
        self.assertEqual(ws.cell(2, build.COL['DPS']).value, '=' + build.L('DPS') + '3')
        self.assertIsNotNone(ws.cell(4, build.COL['DPS']).value)


if __name__ == '__main__':
    unittest.main()
