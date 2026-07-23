import openpyxl
from pathlib import Path

wb_path = Path(r'C:\Users\AedisToru\Documents\GitHub\Cameo-mod\docs\design\cameo_armor_system.xlsx')
wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)

print('Sheets:', wb.sheetnames)

for sheet_name in ['Infantry', 'Tanks', 'Vehicles']:
    ws = wb[sheet_name]
    print(f'\n--- {sheet_name} headers ---')
    header = [ws.cell(row=1, column=c).value for c in range(1, 21)]
    for i, h in enumerate(header, 1):
        print(f'{i}: {h}')
    print('First 5 names:')
    for r in range(2, 7):
        print(r, ws.cell(row=r, column=2).value)

wb.close()
